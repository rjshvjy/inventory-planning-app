"""
output_v2.py — writes the plan Excel for the v2 Inventory Planning App.

Consumes PlanResult (calculation_v2) + Canonical (ingestion_v2) and produces
a workbook by CLONING the master template's 'Appointment plan' sheet shape
(spec §4/§13): per-SKU rows x region columns, priorities row, date stamp —
plus a Calculation_Details sheet carrying every per-line component, flag,
warning, stockout and run-meta item so no adjustment is ever invisible.

Rules honoured:
- Quantities land on each SKU's own master row (master['row']); the writer
  NEVER re-orders or adds SKU rows — the master is the shape contract.
- Manual split regions (e.g. 'Pune (MH)') are left EMPTY (spec §5a: human
  fills by judgment). BLR4 IXD and SellerFlex (YSXA) columns stay empty —
  the plan never ships to the hub or own warehouse.
- Totals (Units Total / Bottles / Litres / Boxes) are computed from the
  written quantities and the master's per-unit columns.
- ASIN-pooled lines (assumption A4) are written on the ACTIVE SKU's row —
  same attribution the engine used.
"""

from __future__ import annotations

import warnings
from datetime import date

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ingestion_v2 import Canonical
from calculation_v2 import PlanResult

warnings.filterwarnings("ignore")

HEADER_ROW = 2
PRIORITY_ROW = 3
DETAILS_SHEET = "Calculation_Details"


def _header_map(ws) -> dict:
    return {str(ws.cell(HEADER_ROW, c).value).strip(): c
            for c in range(1, ws.max_column + 1)
            if ws.cell(HEADER_ROW, c).value}


def build_plan_workbook(res: PlanResult, can: Canonical,
                        template_path, out_path, stock=None,
                        daily_workbook_src=None) -> list[str]:
    """Clone the master template and fill the plan. Returns warnings raised
    during writing (missing columns, unplaceable rows) — caller shows them.

    v2.9.2 combined output: when daily_workbook_src (path or file-like of the
    user's FILLED daily stock workbook) is given, its working tabs are copied
    into the output — one file doubles as inventory plan AND daily stock
    report. Copied tabs keep live formulas; named ranges (LEADTBL/SKUTBL)
    travel too, or the Overall-stock formulas would break. Stock snapshot is
    KEPT (values-only = renders in mobile/email preview, where formula tabs
    show blank until a real spreadsheet app recalculates)."""
    warns: list[str] = []
    wb = load_workbook(template_path)
    # v2.9.1: house style harvested from the template's own Stock snapshot
    from copy import copy as _copy
    _sty = wb["Stock snapshot"].cell(2, 1) if "Stock snapshot" in \
        wb.sheetnames else wb["Appointment plan"].cell(2, 1)
    HDR_FILL, HDR_FONT = _copy(_sty.fill), _copy(_sty.font)
    HDR_BORDER, HDR_ALIGN = _copy(_sty.border), _copy(_sty.alignment)

    def _style_hdr(cell):
        cell.fill, cell.font = _copy(HDR_FILL), _copy(HDR_FONT)
        cell.border, cell.alignment = _copy(HDR_BORDER), _copy(HDR_ALIGN)

    ws = wb["Appointment plan"]
    hdr = _header_map(ws)

    # ---- date stamp: value cell right of the 'From:' label on row 1
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value or "").strip().rstrip(":") == "From":
            ws.cell(1, c + 1, res.meta.get("anchor_date") or str(date.today()))
            break

    # ---- region columns present in the template
    region_col = {r: hdr[r] for r in res.meta.get("region_order", [])
                  if r in hdr}
    for r in res.meta.get("region_order", []):
        if r not in hdr and r not in (can.ixd_region, can.ignore_region):
            warns.append(f"Template has no column for region '{r}' — its "
                         f"quantities were NOT written. Add the column to "
                         f"the master (header row {HEADER_ROW}).")
    manual = set(res.meta.get("manual_split_regions", []))

    # ---- region priorities on row 3
    for region, pri in res.region_priorities.items():
        if region in region_col:
            cell = ws.cell(PRIORITY_ROW, region_col[region], pri)
            cell.font = Font(bold=True)

    # ---- quantities on each SKU's own master row
    sku_row = dict(zip(can.sku_master["sku_u"], can.sku_master["row"]))
    per_unit = can.sku_master.set_index("sku_u")[
        ["units_per_box", "bottles_per_unit", "litres_per_unit"]]
    plan = res.plan[res.plan["rounded_qty"] > 0]
    written: dict[str, float] = {}
    for _, line in plan.iterrows():
        sku, region = line["sku_u"], line["region"]
        if region in manual:
            continue                      # human fills split targets (§5a)
        r = sku_row.get(sku)
        c = region_col.get(region)
        if r is None or c is None:
            warns.append(f"Could not place {sku} x {region} "
                         f"({line['rounded_qty']:.0f}u) on the template.")
            continue
        ws.cell(r, c, int(line["rounded_qty"]))
        written[sku] = written.get(sku, 0) + int(line["rounded_qty"])

    # ---- totals from written quantities + master per-unit data
    for sku, total in written.items():
        r = sku_row[sku]
        pu = per_unit.loc[sku]
        if "Units Total" in hdr:
            ws.cell(r, hdr["Units Total"], int(total))
        if "Bottles" in hdr and pd.notna(pu["bottles_per_unit"]):
            ws.cell(r, hdr["Bottles"], round(total * pu["bottles_per_unit"], 1))
        if "Litres" in hdr and pd.notna(pu["litres_per_unit"]):
            ws.cell(r, hdr["Litres"], round(total * pu["litres_per_unit"], 2))
        if "Boxes" in hdr and pd.notna(pu["units_per_box"]) \
                and pu["units_per_box"]:
            ws.cell(r, hdr["Boxes"], round(total / pu["units_per_box"], 2))

    # ---- Calculation_Details sheet (full transparency)
    if DETAILS_SHEET in wb.sheetnames:
        del wb[DETAILS_SHEET]
    det = wb.create_sheet(DETAILS_SHEET)
    det.cell(1, 1, "CALCULATION DETAILS — every component, flag and warning "
                   "behind the plan. Auto-written; do not edit.")
    det.cell(1, 1).font = Font(bold=True)
    for col_, w_ in (("A", 14), ("B", 18), ("C", 30), ("D", 15), ("E", 12),
                     ("F", 11), ("G", 12), ("H", 12), ("I", 10), ("J", 12),
                     ("K", 12), ("L", 8), ("M", 11), ("N", 13), ("O", 9),
                     ("P", 46)):
        det.column_dimensions[col_].width = w_
    det.freeze_panes = "D5"
    row = 3

    def block(title, df, cols=None):
        nonlocal row
        det.cell(row, 1, title).font = Font(bold=True)
        row += 1
        if df is None or len(df) == 0:
            det.cell(row, 1, "(none)")
            row += 2
            return
        use = cols or list(df.columns)
        for j, cname in enumerate(use, 1):
            _style_hdr(det.cell(row, j, cname))
        det.row_dimensions[row].height = 44   # wrapped headers
        row += 1
        band_on, prev_key = False, object()
        BAND = PatternFill("solid", fgColor="F2F2F2")
        for _, r_ in df.iterrows():
            key = r_.get("ASIN", r_.get(use[0]))
            if key != prev_key:
                band_on, prev_key = not band_on, key
            det.row_dimensions[row].height = 18
            for j, cname in enumerate(use, 1):
                v = r_[cname]
                c_ = det.cell(row, j,
                              round(v, 2) if isinstance(v, float) else v)
                if isinstance(v, float):
                    c_.number_format = "0.00"
                if band_on:
                    c_.fill = PatternFill("solid", fgColor="F2F2F2")
                if cname == "Flags":
                    c_.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        row += 1

    alias = dict(zip(can.sku_master["sku_u"],
                     can.sku_master["item"].where(
                         can.sku_master["item"].astype(bool),
                         can.sku_master["alias"])))
    lead_days = res.meta.get("lead_days", {})
    cover = float(res.meta.get("days_of_cover", 0))
    dp = res.plan.copy()
    dp.insert(1, "product", dp["sku_u"].map(alias))
    dp["target_days"] = dp["region"].map(lead_days).astype(float) + cover
    dp = dp.rename(columns={
        "sku_u": "SKU", "product": "Product", "asin": "ASIN",
        "region": "Region", "daily": "Daily sales (u/day)",
        "current": "Current stock", "in_transit_counted":
        "In-transit counted (in horizon)", "at_fc_pending":
        "At-FC pending (counted)", "ixd_offset": "IXD offset",
        "raw_requirement": "Raw requirement",
        "rounded_qty": "Planned qty (rounded)", "boxes": "Boxes",
        "days_cover_achieved":
        "Days cover AFTER plan (existing+incoming+planned)",
        "target_days": "Target days (lead + cover setting)",
        "priority": "Priority", "flags": "Flags"})
    dp = dp[["ASIN", "SKU", "Product", "Region", "Daily sales (u/day)",
             "Current stock", "In-transit counted (in horizon)",
             "At-FC pending (counted)", "IXD offset", "Raw requirement",
             "Planned qty (rounded)", "Boxes",
             "Target days (lead + cover setting)",
             "Days cover AFTER plan (existing+incoming+planned)",
             "Priority", "Flags"]]
    block("PLAN LINES (days-cover exceeds target when carton rounding "
          "overshoots or existing stock already covers more)", dp)
    block("STOCKOUT WARNINGS (on the rounded plan)", res.stockouts)
    block("RUN WARNINGS",
          pd.DataFrame({"warning": res.warnings})
          if res.warnings else pd.DataFrame())
    meta_df = pd.DataFrame([{"key": k, "value": str(v)}
                            for k, v in res.meta.items()])
    block("RUN META (incl. flagged assumptions)", meta_df)

    # ---- Sales-velocity tab (v2.9 change 4): the "eyes open" view — the
    # rate that drove each region's plan, beside the plan itself.
    VEL_SHEET = "Sales velocity"
    if VEL_SHEET in wb.sheetnames:
        del wb[VEL_SHEET]
    vs = wb.create_sheet(VEL_SHEET)
    vs.cell(1, 1, "SALES VELOCITY vs PLAN — daily rate per region beside the "
                  "planned quantity it produced. Auto-written; do not edit.")
    vs.cell(1, 1).font = Font(bold=True)
    regions = [r for r in res.meta.get("region_order", [])
               if r in set(res.plan["region"])]
    pq = res.plan.pivot_table(index="sku_u", columns="region",
                              values="rounded_qty", aggfunc="sum",
                              fill_value=0)
    dv_ = res.plan.pivot_table(index="sku_u", columns="region",
                               values="daily", aggfunc="sum", fill_value=0.0)
    import numpy as _np
    _dcsrc = res.plan.replace([_np.inf, -_np.inf], _np.nan)
    dc_ = _dcsrc.pivot_table(index="sku_u", columns="region",
                             values="days_cover_achieved", aggfunc="mean")
    COVER = float(res.meta.get("days_of_cover", 0))
    AMBER = PatternFill("solid", fgColor="FFE699")
    sku_order = [s_ for s_ in can.sku_master["sku_u"] if s_ in pq.index]
    # header rows 3 (region, merged over 2 cols) and 4 (Plan | Velocity)
    sku_asin = dict(zip(can.sku_master["sku_u"], can.sku_master["asin"]))
    for c0, t0 in ((1, "ASIN"), (2, "SKU"), (3, "Product")):
        _style_hdr(vs.cell(3, c0, t0)); _style_hdr(vs.cell(4, c0, ""))
    col = 4
    SUB = ("Plan qty", "Velocity /day", "Days cover")
    for rg in regions + ["TOTAL"]:
        _style_hdr(vs.cell(3, col, rg))
        vs.merge_cells(start_row=3, start_column=col,
                       end_row=3, end_column=col + 2)
        for k, t_ in enumerate(SUB):
            _style_hdr(vs.cell(4, col + k, t_))
        col += 3
    vs.row_dimensions[3].height = 18
    vs.row_dimensions[4].height = 28
    REG_BORDER = Border(left=Side(style="medium", color="1F4E5F"))
    band_starts = list(range(4, col, 3))

    def _trio(r0, c0, q, v, dc=None):
        if q:
            vs.cell(r0, c0, int(q))
        if v:
            vs.cell(r0, c0 + 1, round(v, 2)).number_format = "0.00"
            if dc is None:
                dc = q / v if q else None
            if dc is not None and dc == dc:      # not NaN
                c2 = vs.cell(r0, c0 + 2, round(float(dc), 1))
                c2.number_format = "0.0"
                if COVER and dc > COVER * 1.5:   # 50% slack: carton rounding is not overstock
                    c2.fill = AMBER              # over the cover setting
    r_ = 5
    BAND = PatternFill("solid", fgColor="F2F2F2")
    for i_, sku in enumerate(sku_order):
        vs.cell(r_, 1, sku_asin.get(sku, ""))
        vs.cell(r_, 2, sku)
        vs.cell(r_, 3, alias.get(sku, ""))
        vs.row_dimensions[r_].height = 18
        col = 4
        wsum, vsum = 0.0, 0.0
        for rg in regions:
            v0 = float(dv_.loc[sku].get(rg, 0.0))
            d0 = (float(dc_.loc[sku].get(rg))
                  if rg in dc_.columns and dc_.loc[sku].notna().get(rg, False)
                  else None)
            _trio(r_, col, int(pq.loc[sku].get(rg, 0)), v0, dc=d0)
            if v0 and d0 is not None:
                wsum += v0 * d0; vsum += v0
            col += 3
        _trio(r_, col, int(pq.loc[sku].sum()), float(dv_.loc[sku].sum()),
              dc=(wsum / vsum if vsum else None))
        if i_ % 2:
            for c_ in range(1, col + 3):
                vs.cell(r_, c_).fill = PatternFill("solid",
                                                   fgColor="F2F2F2")
        r_ += 1
    vs.cell(r_, 1, "TOTAL")
    vs.row_dimensions[r_].height = 18
    col = 4
    gw, gv = 0.0, 0.0
    for rg in regions:
        w_ = float((dv_[rg] * dc_[rg].fillna(0)).sum()) \
            if rg in dc_.columns else 0.0
        v_ = float(dv_[rg][dc_[rg].notna()].sum()) \
            if rg in dc_.columns else 0.0
        _trio(r_, col, int(pq[rg].sum()), float(dv_[rg].sum()),
              dc=(w_ / v_ if v_ else None))
        gw += w_; gv += v_
        col += 3
    _trio(r_, col, int(pq.values.sum()), float(dv_.values.sum()),
          dc=(gw / gv if gv else None))
    for c_ in range(1, col + 3):
        vs.cell(r_, c_).font = Font(bold=True)
    # medium teal border at every region-block start, header to TOTAL
    for c_ in band_starts + [col]:
        for rr in range(3, r_ + 1):
            cell = vs.cell(rr, c_)
            cell.border = cell.border + REG_BORDER if cell.border else \
                REG_BORDER
    vs.column_dimensions["A"].width = 14
    vs.column_dimensions["B"].width = 20
    vs.column_dimensions["C"].width = 30
    from openpyxl.utils import get_column_letter
    for c_ in range(4, col + 3):
        vs.column_dimensions[get_column_letter(c_)].width = 11
    vs.freeze_panes = "D5"

    # ---- Stock snapshot (v2.9.1): values from the uploaded workbook's
    # current stock, written onto the template's own SKU rows.
    if stock is not None and "Stock snapshot" in wb.sheetnames:
        ss = wb["Stock snapshot"]
        ss_hdr = {str(ss.cell(2, c).value).strip(): c
                  for c in range(1, ss.max_column + 1) if ss.cell(2, c).value}
        import pandas as _pd
        cur = stock.current.pivot_table(index="sku_u", columns="region",
                                        values="qty", aggfunc="sum",
                                        fill_value=0)
        itp = (stock.in_transit.pivot_table(index="sku_u", columns="region",
                                            values="qty", aggfunc="sum",
                                            fill_value=0)
               if stock.in_transit is not None and len(stock.in_transit)
               else _pd.DataFrame())
        afp = (stock.at_fc.pivot_table(index="sku_u", columns="region",
                                       values="pending", aggfunc="sum",
                                       fill_value=0)
               if stock.at_fc is not None and len(stock.at_fc)
               else _pd.DataFrame())
        sku_col = ss_hdr.get("SKU", 1)
        lbl_col = ss_hdr.get("Row Labels", 2)
        for r_ in range(3, ss.max_row + 1):
            sku = ss.cell(r_, sku_col).value
            if not sku:
                continue
            sku_u = str(sku).strip().upper()
            lbl = str(ss.cell(r_, lbl_col).value or "")
            src = (itp if lbl.startswith("In transit") else
                   afp if lbl.startswith("At the FC") else cur)
            if sku_u not in getattr(src, "index", []):
                continue
            total = 0
            for region, c_ in ss_hdr.items():
                if region in src.columns:
                    v = int(src.loc[sku_u, region])
                    if v:
                        ss.cell(r_, c_, v)
                    total += v
            if "Grand Total" in ss_hdr and total:
                ss.cell(r_, ss_hdr["Grand Total"], total)
        if getattr(stock, "stock_as_of", None) is not None:
            ss.cell(1, ss.max_column,
                    "As of " + stock.stock_as_of.strftime("%d-%b-%Y"))

    # ---- v2.9.2: combined file — graft the daily workbook's tabs in
    if daily_workbook_src is not None:
        try:
            _merge_daily_tabs(wb, daily_workbook_src)
            _write_combined_readme(wb)
        except Exception as e:                       # never sink the plan
            warns.append(f"Combined output: daily tabs could not be copied "
                         f"({e}) — plan tabs are complete; the daily "
                         f"workbook remains your separate file this run.")

    # ---- v2.9.2: tab colours + reading order (names unchanged — formulas
    #      and the reader reference tabs by name, so we NEVER rename)
    AMBER, TEAL, GREY = "FFC000", "1F4E5F", "BFBFBF"   # amber = act on
    colour = {"README": GREY,
              "Appointment plan": AMBER, DETAILS_SHEET: AMBER,
              "Sales velocity": AMBER,
              "Stock snapshot": TEAL, "Reference": TEAL,
              "Current stock": TEAL, "In-transit": TEAL,
              "At the FC": TEAL, "Overall stock": TEAL,
              "EDD & Expiry": TEAL}
    for name, col in colour.items():
        if name in wb.sheetnames:
            wb[name].sheet_properties.tabColor = col
    ORDER = ["README", "Appointment plan", DETAILS_SHEET, "Sales velocity",
             "Stock snapshot", "Reference", "Current stock", "In-transit",
             "At the FC", "Overall stock", "EDD & Expiry"]
    rank = {t: i for i, t in enumerate(ORDER)}
    wb._sheets.sort(key=lambda ws_: rank.get(ws_.title, len(ORDER)))
    wb.active = 0

    wb.save(out_path)
    return warns


DAILY_TABS = ["Reference", "Current stock", "In-transit", "At the FC",
              "Overall stock", "EDD & Expiry"]      # all-or-nothing family


def _merge_daily_tabs(wb, src):
    """Copy the daily workbook's working tabs (values, formulas, styles,
    widths, merges) plus the named ranges their formulas need."""
    from copy import copy as _c
    daily = load_workbook(src)
    for t in DAILY_TABS:
        if t not in daily.sheetnames:
            continue
        if t in wb.sheetnames:                       # avoid collision
            del wb[t]
        src_ws, ns = daily[t], wb.create_sheet(t)
        for row_ in src_ws.iter_rows():
            for cell in row_:
                nc = ns.cell(cell.row, cell.column, cell.value)
                if cell.has_style:
                    nc.font, nc.fill = _c(cell.font), _c(cell.fill)
                    nc.border = _c(cell.border)
                    nc.alignment = _c(cell.alignment)
                    nc.number_format = cell.number_format
        for col_, dim in src_ws.column_dimensions.items():
            ns.column_dimensions[col_].width = dim.width
        for r_, dim in src_ws.row_dimensions.items():
            ns.row_dimensions[r_].height = dim.height
        for mc in src_ws.merged_cells.ranges:
            ns.merge_cells(str(mc))
        ns.freeze_panes = src_ws.freeze_panes
    for name, defn in daily.defined_names.items():   # LEADTBL / SKUTBL
        if name not in wb.defined_names:
            wb.defined_names[name] = defn


def _write_combined_readme(wb):
    """Overwrite the README with the dual-purpose orientation."""
    from openpyxl.styles import Font as _F
    if "README" not in wb.sheetnames:
        return
    rd = wb["README"]
    for r_ in range(1, 45):
        for c_ in range(1, 8):
            rd.cell(r_, c_).value = None
    lines = [
        ("COMBINED INVENTORY FILE — plan + daily stock report in one "
         "workbook", 1),
        ("", 0),
        ("AMBER tabs = THE PLAN (act on these):", 1),
        ("  Appointment plan — units to ship per product per region, with "
         "priorities. THE deliverable.", 0),
        ("  Calculation_Details — every number, flag (incl. IXD_ASSUMED) "
         "and warning behind the plan.", 0),
        ("  Sales velocity — plan vs daily sales vs days-cover; amber "
         "cells = possible overstock.", 0),
        ("", 0),
        ("TEAL tabs = DAILY STOCK REPORT (keep these current):", 1),
        ("  Stock snapshot — read-only stock values (this one shows in "
         "phone/email previews).", 0),
        ("  Reference — lead times + SKU list; correct lead times when "
         "they change.", 0),
        ("  Current stock — on-hand per region (green = prefilled from "
         "the ledger).", 0),
        ("  In-transit — shipments on the way; use the dropdowns.", 0),
        ("  At the FC — arrivals being received; enter Shipped & "
         "Received.", 0),
        ("  Overall stock — auto-built summary. FORMULAS — do not type "
         "over; may show blank in phone previews until opened in Excel.",
         0),
        ("  EDD & Expiry — your delivery/expiry scratch tab.", 0),
        ("", 0),
        ("The daily tabs are a SNAPSHOT from when this file was "
         "generated — regenerate for fresh data.", 0),
        ("Dates display Indian-style (07-Jul-2026).", 0),
    ]
    for i, (txt, bold) in enumerate(lines, start=1):
        cell = rd.cell(i, 1, txt)
        cell.font = _F(bold=bool(bold),
                       size=13 if i == 1 else 11,
                       color="1F4E5F" if bold else "1A2226")
    rd.column_dimensions["A"].width = 100
