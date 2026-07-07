"""
workbook_builder_v2.py — Daily stock workbook: fresh-build writer + reader.

Owns the stock-workbook SCHEMA for BOTH directions (spec option (a), v2.6
fresh-build-by-code): tab names, column layouts, named ranges, dropdowns,
formulas, and formats are defined once here, so the file the app hands out
and the file it reads back can never drift apart.

Writer  : build_stock_workbook(canonical, prefill=True)  -> bytes (xlsx)
Reader  : read_stock_workbook(file, canonical=None)      -> (StockInputs, Report)

Spec refs:
  §7a   input tabs (Current stock / In-transit / At the FC), derived
        Overall-stock report the app NEVER reads, lead-days table in workbook
  §7a-2 EDD & Expiry human-only scaffold; fresh build = totals computed from
        the live row count
  v2.5  CSV-driven pre-fill (Option A): Current stock from ledger (per region)
        + General CSV as-on-date column; In-transit/At-FC stay skeletons
  v2.5  IXD counted in grand total, excluded from "Regional (excl IXD & YSXA)"

Reads canonical data from ingestion_v2 only — never raw Amazon files.
"""

from __future__ import annotations
import io
from dataclasses import dataclass, field
from datetime import datetime

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from ingestion_v2 import Report, Canonical, _to_date

# ================================================================== SCHEMA
# Single source of truth for tab names, special columns, and row capacities.

TAB_README = "README"
TAB_REFERENCE = "Reference"
TAB_CURRENT = "Current stock"
TAB_INTRANSIT = "In-transit"
TAB_ATFC = "At the FC"
TAB_OVERALL = "Overall stock"
TAB_EDD = "EDD & Expiry"

COL_IXD = "BLR4 IXD"           # grand-total on-hand, valid destination
COL_YSXA = "YSXA"              # visible on stock tabs, never a destination

INTRANSIT_ROWS = 60            # blank input rows provided (3..62)
ATFC_ROWS = 40                 # blank input rows provided (3..42)

MODES = ("ATS", "Self")

# Seed lead-days for a fresh build (user-editable per run; Reference tab is
# the runtime home of lead times, spec §7a "where facts live").
DEFAULT_LEAD_DAYS = {
    "Delhi (DL)": 13, "Haryana (HR)": 13, "Bombay (MH)": 10, "Pune (MH)": 10,
    "Bangalore (KA)": 3, "Hyderabad (TG)": 3, "Chennai (TN)": 2,
    "Calcutta (WB)": 14, "Lucknow (UP)": 12, COL_IXD: 3,
}
FALLBACK_LEAD_DAYS = 7         # seed for a region with no known default

# ------------------------------------------------------------------ styles
ARIAL10 = Font(name="Arial", size=10)
HDR_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
HDR_FILL = PatternFill("solid", start_color="FF1F4E5F")
GOLD_FONT = Font(name="Arial", size=10, bold=True, color="FF7F6000")
GOLD_HDR_FILL = PatternFill("solid", start_color="FF7F6000")
GOLD_FILL = PatternFill("solid", start_color="FFFFF2CC")
YELLOW_FILL = PatternFill("solid", start_color="FFFFFF00")
INPUT_FONT = Font(name="Arial", size=10, color="FF0000FF")       # human-typed
PREFILL_FONT = Font(name="Arial", size=10, color="FF006100")     # app-filled
AUTO_FONT = Font(name="Arial", size=10, color="FF006100")        # formulas
NOTE_FONT = Font(name="Arial", size=9, italic=True, color="FF808080")
TITLE_FONT = Font(name="Arial", size=10, bold=True)
SKU_FONT = Font(name="Arial", size=10, bold=True)
WARN_FONT = Font(name="Arial", size=11, bold=True, color="FF9C0006")
THIN = Side(style="thin", color="FFC9C9C9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GREY_FILL = PatternFill("solid", start_color="FFF2F2F2")
DATE_FMT = "dd\\-mmm\\-yy"
NUM_FMT = "#,##0"
NUM_FMT_BLANKZERO = '#,##0;\\-#,##0;""'

# Overall-stock banding: alternating blue/grey SKU blocks, darkest Available.
BAND_BLUE = ("FFBDD7EE", "FFDDEBF7", "FFEDF4FB")
BAND_GREY = ("FFD9D9D9", "FFE7E7E7", "FFF2F2F2")

# CF palette (matches delivered file)
CF_RED = "FFFFC7CE"
CF_AMBER = "FFFFEB9C"
CF_GREEN = "FFC6EFCE"
OIL_TINTS = [  # (search term, fill) — order matters: first match wins
    ("virgin coconut", "FFD9E1F2"), ("coconut", "FFC6EFCE"),
    ("sesame", "FFFFF2CC"), ("groundnut", "FFFCE4D6"),
    ("mustard", "FFEDEDED"), ("deepam", "FFE4DFEC"),
]

README_LINES = [
    ("DAILY STOCK WORKBOOK — READ ME", True),
    ("This workbook is BOTH the app's stock input AND your daily inventory report. The app builds it", False),
    ("FRESH on every download (SKUs from the master, regions from the FC registration — so a new SKU or", False),
    ("newly registered state appears automatically) and PRE-FILLS the green cells from your two Amazon", False),
    ("stock CSVs. You complete the blue cells, glance at the report, upload. One sitting.", False),
    ("TABS, IN YOUR WORKING ORDER", True),
    ("1. CURRENT STOCK — app-filled: per-region on-hand from the FC-wise ledger (2-day-old; SELLABLE", False),
    ("   only; YSXA & BLR4 IXD separate) PLUS the yellow 'Current stock (as-on-date)' column = TRUE", False),
    ("   national fulfillable today (General CSV). They differ BY DESIGN: locations say WHERE (2 days", False),
    ("   ago), as-on-date says HOW MUCH (today). Set the 'Stock as of' date to the Amazon data's date.", False),
    ("2. IN-TRANSIT — one row per SKU per shipment; REPEAT the Shipment ID every row. Dropdowns for", False),
    ("   Destination (regions incl. Pune (MH) and BLR4 IXD), Mode, SKU. Enter Quantity + Cargo ready", False),
    ("   date. ATS: suggestion = cargo + region lead days (Reference tab). Self: suggestion = cargo date", False),
    ("   (you ship early to hit the appointment). Suggestion is ADVICE — type the committed date in", False),
    ("   'Actual reach date'; it stays RED until you do and the app rejects the file otherwise.", False),
    ("3. AT THE FC — arrived, being checked in. Shipped & Received AS NUMBERS ('fully received' =", False),
    ("   Received=Shipped). Pending computes itself; amber = Received still blank.", False),
    ("4. OVERALL STOCK — auto-built report, protected, never type here: three rows per SKU across all", False),
    ("   regions + Grand Total (everything incl. YSXA), Regional + IXD (excl. YSXA) — IXD counts because", False),
    ("   that stock lands in some FC within days — and Current stock (as-on-date).", False),
    ("5. EDD & EXPIRY — human-only; the app never reads it. Set the Report date, enter each region's", False),
    ("   estimated Delivery Date (from Amazon PDP checks); Days and Avg compute and colour: green <2,", False),
    ("   amber 2-3, red >3. Enter each SKU's Best-before date: RED = already expired, AMBER = within 60", False),
    ("   days of the report date. Product names tint by oil family automatically.", False),
    ("REFERENCE TAB", True),
    ("Region lead-days (blue) feed reach-date suggestions — tweak when transport reality changes; the", False),
    ("app echoes values used in its output Summary. SKU list mirrors the master (regenerated each build).", False),
    ("RULES", True),
    ("Complete + upload in ONE session (pre-filled data carries its date; stale gets flagged). Don't", False),
    ("rename sheets/headers. Clear a row's contents to remove an entry; no half-filled rows. Same SKU", False),
    ("repeating within one shipment is fine (different carton configs). YSXA is never a destination.", False),
    ("SKU ORDER & OVERALL-STOCK BANDING", True),
    ("All tabs share one canonical SKU order (oil groups by 3-month sales, sizes descending, Temps glued", False),
    ("to their base) — identical to the plan template, so cross-file verification is row-for-row. In the", False),
    ("Overall stock report each SKU's three rows form one colour block (alternating blue/grey), shaded", False),
    ("darkest on Available, lighter on In transit, lightest on At the FC.", False),
]


# =============================================================== helpers

def ordered_regions(can: Canonical) -> list[str]:
    """canonical.regions with split-state regions glued together
    (non-default immediately after its state's default), preserving the
    registration-file order otherwise."""
    regions = list(can.regions)
    for st in {s["state"] for s in can.region_splits}:
        rows = [s for s in can.region_splits if s["state"] == st]
        default = next((s for s in rows if s["default"]), None)
        if not default:
            continue
        dft_lbl = f"{default['label']} ({st})"
        others = [f"{s['label']} ({st})" for s in rows if not s["default"]]
        others = [o for o in others if o in regions]
        if dft_lbl not in regions:
            continue
        for o in others:
            regions.remove(o)
        i = regions.index(dft_lbl)
        for j, o in enumerate(others, 1):
            regions.insert(i + j, o)
    return regions


def _lead_seed(region: str) -> int:
    return DEFAULT_LEAD_DAYS.get(region, FALLBACK_LEAD_DAYS)


def _set(ws, r, c, value, font=None, fill=None, fmt=None, border=True,
         align=None):
    cell = ws.cell(r, c, value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if border:
        cell.border = BORDER
    if align:
        cell.alignment = align
    return cell


def _hdr(ws, r, c, value, fill=HDR_FILL, wrap=True, valign="center"):
    return _set(ws, r, c, value, font=HDR_FONT if fill == HDR_FILL else
                Font(name="Arial", size=10, bold=True, color="FFFFFFFF"),
                fill=fill,
                align=Alignment(horizontal="center", vertical=valign,
                                wrap_text=wrap))


def _cf(ws, rng, formula, bg, stop=False):
    dxf = DifferentialStyle(fill=PatternFill(start_color=bg, end_color=bg,
                                             fill_type="solid"))
    rule = Rule(type="expression", formula=[formula], dxf=dxf,
                stopIfTrue=stop or None)
    ws.conditional_formatting.add(rng, rule)


# ================================================================= WRITER

def build_stock_workbook(can: Canonical, prefill: bool = True,
                         today: pd.Timestamp | None = None) -> bytes:
    """Fresh-build the daily stock workbook. prefill=True fills the
    Current-stock tab from canonical.available (per region) and the
    as-on-date column from canonical.national (Option A); False emits the
    blank downloadable template. Returns xlsx bytes."""
    today = today or pd.Timestamp.now().normalize()
    regions = ordered_regions(can)
    stock_cols = regions + [COL_IXD, COL_YSXA]      # Current / Overall tabs
    dest_list = regions + [COL_IXD]                  # dropdowns + lead table
    master = can.sku_master
    skus = list(master["sku"])
    products = list(master["item"]) if "item" in master else [""] * len(skus)
    n = len(skus)

    wb = Workbook()
    wb.remove(wb.active)

    # ---------------------------------------------------------- README
    ws = wb.create_sheet(TAB_README)
    ws.column_dimensions["A"].width = 110
    for i, (line, bold) in enumerate(README_LINES, 1):
        c = ws.cell(i, 1, line)
        c.font = Font(name="Arial", size=12 if i == 1 else 10, bold=bold)

    # ------------------------------------------------------- Reference
    ws = wb.create_sheet(TAB_REFERENCE)
    for col, w in (("A", 16), ("B", 10), ("D", 22), ("E", 36)):
        ws.column_dimensions[col].width = w
    ws.cell(1, 1, "LEAD TIMES (days, factory -> region) — edit per run if "
                  "reality has changed").font = TITLE_FONT
    ws.cell(1, 4, "SKU LIST (feeds dropdowns; keep in sync with master "
                  "template)").font = TITLE_FONT
    for c, t in ((1, "Region"), (2, "Lead days"), (4, "SKU"), (5, "Product")):
        _set(ws, 2, c, t, font=HDR_FONT, fill=HDR_FILL)
    for i, rg in enumerate(dest_list, 3):
        _set(ws, i, 1, rg)
        _set(ws, i, 2, _lead_seed(rg), font=INPUT_FONT)
    for i, (sku, prod) in enumerate(zip(skus, products), 3):
        _set(ws, i, 4, sku)
        _set(ws, i, 5, prod)
    lead_end = 2 + len(dest_list)
    sku_end = 2 + n
    wb.defined_names.add(DefinedName(
        "LEADTBL", attr_text=f"{TAB_REFERENCE}!$A$3:$B${lead_end}"))
    wb.defined_names.add(DefinedName(
        "SKUTBL", attr_text=f"{TAB_REFERENCE}!$D$3:$E${sku_end}"))

    # --------------------------------------------------- Current stock
    ws = wb.create_sheet(TAB_CURRENT)
    ws.freeze_panes = "C3"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 12
    ncols = 2 + len(stock_cols)            # A,B + regions/IXD/YSXA
    total_col = ncols + 1
    asod_col = ncols + 2
    ws.column_dimensions[get_column_letter(asod_col)].width = 15

    ws.cell(1, 1, "Stock as of (date):").font = TITLE_FONT
    d = ws.cell(1, 2)
    d.fill, d.number_format = YELLOW_FILL, DATE_FMT
    note = ws.cell(1, 4, "<- date OF THE AMAZON DATA (~2 days behind live). "
                         "App pre-fills green cells from your CSVs.")
    note.font = NOTE_FONT
    _hdr(ws, 2, 1, "SKU", valign="bottom")
    _hdr(ws, 2, 2, "Product", valign="bottom")
    for i, rg in enumerate(stock_cols, 3):
        _hdr(ws, 2, i, rg, valign="bottom")
    _hdr(ws, 2, total_col, "Total", valign="bottom")
    _hdr(ws, 2, asod_col, "Current stock (as-on-date)", fill=GOLD_HDR_FILL,
         valign="bottom")
    ws.row_dimensions[2].height = 42

    avail = can.available if can.available is not None else pd.DataFrame(
        columns=["sku_u", "region", "bal"])
    piv = (avail.pivot_table(index="sku_u", columns="region", values="bal",
                             aggfunc="sum").fillna(0)
           if len(avail) else pd.DataFrame())
    nat = (can.national.set_index("sku_u")["fulfillable"]
           if prefill and can.national is not None else pd.Series(dtype=float))

    ledger_date = can.effective_dates.get("ledger")
    if prefill and ledger_date is not None:
        d.value = ledger_date

    for i, (sku, prod) in enumerate(zip(skus, products)):
        r = 3 + i
        sku_u = sku.upper()
        _set(ws, r, 1, sku, font=SKU_FONT)
        _set(ws, r, 2, prod)
        for j, rg in enumerate(stock_cols, 3):
            v = None
            if prefill and len(piv) and \
                    sku_u in piv.index and rg in piv.columns:
                pv = float(piv.at[sku_u, rg])
                v = int(pv) if pv else None
            _set(ws, r, j, v, font=PREFILL_FONT if prefill else INPUT_FONT,
                 fmt=NUM_FMT)
        L = get_column_letter(2 + len(stock_cols))
        _set(ws, r, total_col, f"=SUM(C{r}:{L}{r})", fmt=NUM_FMT)
        av = None
        if prefill and sku_u in nat.index:
            nv = float(nat.loc[sku_u])
            av = int(nv)
        _set(ws, r, asod_col, av, font=GOLD_FONT, fill=GOLD_FILL, fmt=NUM_FMT)

    # ------------------------------------------------------ In-transit
    ws = wb.create_sheet(TAB_INTRANSIT)
    ws.freeze_panes = "A3"
    for col, w in (("A", 16), ("B", 17), ("C", 8), ("D", 20), ("E", 32),
                   ("F", 10), ("G", 14), ("H", 16), ("I", 16)):
        ws.column_dimensions[col].width = w
    ws.cell(1, 1, "Generated for planning date:").font = TITLE_FONT
    d = ws.cell(1, 2, today)
    d.fill, d.number_format = YELLOW_FILL, DATE_FMT
    ws.cell(1, 3, "One row per SKU per shipment. REPEAT the Shipment ID on "
                  "every row (no blanks).").font = NOTE_FONT
    it_headers = ["Shipment ID", "Destination (region)", "Mode", "SKU",
                  "Product (auto)", "Quantity", "Cargo ready date",
                  "Suggested reach date (auto)", "Actual reach date (YOU fill)"]
    for c, t in enumerate(it_headers, 1):
        _hdr(ws, 2, c, t)
    ws.row_dimensions[2].height = 42
    it_last = 2 + INTRANSIT_ROWS
    for r in range(3, it_last + 1):
        for c, fmt in ((1, None), (2, None), (3, None), (4, None)):
            _set(ws, r, c, None, font=INPUT_FONT, fmt=fmt)
        _set(ws, r, 5, f'=IFERROR(VLOOKUP($D{r},SKUTBL,2,0),"")',
             font=AUTO_FONT)
        _set(ws, r, 6, None, font=INPUT_FONT, fmt=NUM_FMT)
        _set(ws, r, 7, None, font=INPUT_FONT, fmt=DATE_FMT)
        _set(ws, r, 8, f'=IF(OR($D{r}="",$G{r}="",$C{r}="",$B{r}=""),"",'
                       f'IF($C{r}="Self",$G{r},$G{r}+IFERROR('
                       f'VLOOKUP($B{r},LEADTBL,2,0),0)))',
             font=AUTO_FONT, fill=GREY_FILL, fmt=DATE_FMT)
        _set(ws, r, 9, None, font=INPUT_FONT, fmt=DATE_FMT)
    dest_csv = ",".join(dest_list)
    dv_dest = DataValidation(type="list", formula1=f'"{dest_csv}"',
                             allow_blank=True)
    dv_mode = DataValidation(type="list", formula1=f'"{",".join(MODES)}"',
                             allow_blank=True)
    dv_sku = DataValidation(type="list",
                            formula1=f"{TAB_REFERENCE}!$D$3:$D${sku_end}",
                            allow_blank=True)
    for dv, col in ((dv_dest, "B"), (dv_mode, "C"), (dv_sku, "D")):
        ws.add_data_validation(dv)
        dv.add(f"{col}3:{col}{it_last}")
    _cf(ws, f"I3:I{it_last}", f'AND($D3<>"",$I3="")', CF_RED)

    # ------------------------------------------------------- At the FC
    ws = wb.create_sheet(TAB_ATFC)
    ws.freeze_panes = "A3"
    for col, w in (("A", 16), ("B", 17), ("C", 20), ("D", 32), ("E", 11),
                   ("F", 11), ("G", 14)):
        ws.column_dimensions[col].width = w
    ws.cell(1, 1, "Shipments that have reached the FC and are being received."
                  " App pre-fills from Amazon where possible.").font = \
        TITLE_FONT
    fc_headers = ["Shipment ID", "Destination (region)", "SKU",
                  "Product (auto)", "Shipped Qty", "Received Qty",
                  "Units at FC pending (auto)", "Cargo ready date",
                  "Receiving date"]
    for c, t in enumerate(fc_headers, 1):
        _hdr(ws, 2, c, t)
    ws.row_dimensions[2].height = 42
    fc_last = 2 + ATFC_ROWS
    for r in range(3, fc_last + 1):
        for c in (1, 2, 3):
            _set(ws, r, c, None, font=INPUT_FONT)
        _set(ws, r, 4, f'=IFERROR(VLOOKUP($C{r},SKUTBL,2,0),"")',
             font=AUTO_FONT)
        _set(ws, r, 5, None, font=INPUT_FONT, fmt=NUM_FMT)
        _set(ws, r, 6, None, font=INPUT_FONT, fmt=NUM_FMT)
        _set(ws, r, 7, f'=IF(OR($E{r}="",$F{r}=""),"",$E{r}-$F{r})',
             fill=GREY_FILL, fmt=NUM_FMT)
        _set(ws, r, 8, None, font=INPUT_FONT, fmt=DATE_FMT)
        _set(ws, r, 9, None, font=INPUT_FONT, fmt=DATE_FMT)
    dv_dest2 = DataValidation(type="list", formula1=f'"{dest_csv}"',
                              allow_blank=True)
    dv_sku2 = DataValidation(type="list",
                             formula1=f"{TAB_REFERENCE}!$D$3:$D${sku_end}",
                             allow_blank=True)
    ws.add_data_validation(dv_dest2)
    dv_dest2.add(f"B3:B{fc_last}")
    ws.add_data_validation(dv_sku2)
    dv_sku2.add(f"C3:C{fc_last}")
    _cf(ws, f"F3:F{fc_last}", f'AND($E3<>"",$F3="")', CF_AMBER)

    # ---------------------------------------------------- Overall stock
    ws = wb.create_sheet(TAB_OVERALL)
    ws.freeze_panes = "C3"
    ws.protection.sheet = True
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12
    o_ncols = 2 + len(stock_cols)
    gt_col = o_ncols + 1                    # Grand Total
    reg_col = o_ncols + 2                   # Regional (excl IXD & YSXA)
    typ_col = o_ncols + 3                   # Row type
    asod_o = o_ncols + 4                    # Current stock (as-on-date)
    ws.column_dimensions[get_column_letter(asod_o)].width = 15
    ws.cell(1, 1, "OVERALL STOCK — auto-built daily report. DO NOT TYPE "
                  "HERE; it rebuilds from the input tabs.").font = WARN_FONT
    _hdr(ws, 2, 1, "SKU")
    _hdr(ws, 2, 2, "Row Labels")
    for i, rg in enumerate(stock_cols, 3):
        _hdr(ws, 2, i, rg)
    _hdr(ws, 2, gt_col, "Grand Total")
    _hdr(ws, 2, reg_col, "Regional + IXD (excl. YSXA)")
    _hdr(ws, 2, typ_col, "Row type")
    _hdr(ws, 2, asod_o, "Current stock (as-on-date)", fill=GOLD_HDR_FILL)
    ws.column_dimensions[get_column_letter(reg_col)].width = 13
    ws.row_dimensions[2].height = 42

    last_stock_L = get_column_letter(2 + len(stock_cols))       # incl. YSXA
    last_reg_L = get_column_letter(2 + len(regions))            # regions only
    ysxa_idx = 2 + len(stock_cols)                              # YSXA col no.
    cur_last = 2 + n
    data_last = 2 + 3 * n

    for i, (sku, prod) in enumerate(zip(skus, products)):
        band = (BAND_BLUE if i % 2 == 0 else BAND_GREY)
        base = 3 + 3 * i
        for k, (label, typ) in enumerate((
                (prod, "Available"),
                (f"In transit - {prod}", "In transit"),
                (f"At the FC - {prod}", "At the FC"))):
            r = base + k
            fill = PatternFill("solid", start_color=band[k])
            a = _set(ws, r, 1, sku,
                     font=SKU_FONT if k == 0 else ARIAL10,
                     fill=fill, border=False)
            _set(ws, r, 2, label, fill=fill, border=False)
            for j, rg in enumerate(stock_cols, 3):
                L = get_column_letter(j)
                if typ == "Available":
                    f = (f"=SUMIF('{TAB_CURRENT}'!$A:$A,$A{r},"
                         f"'{TAB_CURRENT}'!{L}:{L})")
                elif rg == COL_YSXA:
                    f = "=0"
                elif typ == "In transit":
                    f = (f"=SUMIFS('{TAB_INTRANSIT}'!$F:$F,"
                         f"'{TAB_INTRANSIT}'!$D:$D,$A{r},"
                         f"'{TAB_INTRANSIT}'!$B:$B,{L}$2)")
                else:
                    f = (f"=SUMIFS('{TAB_ATFC}'!$G:$G,"
                         f"'{TAB_ATFC}'!$C:$C,$A{r},"
                         f"'{TAB_ATFC}'!$B:$B,{L}$2)")
                _set(ws, r, j, f, fill=fill, fmt=NUM_FMT_BLANKZERO)
            ixd_L = get_column_letter(2 + len(regions) + 1)   # IXD column
            _set(ws, r, gt_col, f"=SUM(C{r}:{last_stock_L}{r})",
                 font=SKU_FONT, fill=fill, fmt=NUM_FMT_BLANKZERO)
            _set(ws, r, reg_col,
                 f"=SUM(C{r}:{last_reg_L}{r})+{ixd_L}{r}",
                 font=SKU_FONT, fill=fill, fmt=NUM_FMT_BLANKZERO)
            _set(ws, r, typ_col, typ, font=NOTE_FONT, border=False)
            if typ == "Available":
                _set(ws, r, asod_o,
                     f"=SUMIF('{TAB_CURRENT}'!$A:$A,$A{r},"
                     f"'{TAB_CURRENT}'!${get_column_letter(asod_col)}:"
                     f"${get_column_letter(asod_col)})",
                     font=GOLD_FONT, fill=GOLD_FILL, fmt=NUM_FMT)

    # footer totals — ranges computed from the LIVE row count (spec v2.6)
    for k, typ in enumerate(("Available", "In transit", "At the FC")):
        r = data_last + 1 + k
        _set(ws, r, 2, f"TOTAL {typ}", font=SKU_FONT)
        for j in range(3, reg_col + 1):
            L = get_column_letter(j)
            _set(ws, r, j,
                 f'=SUMIF($'
                 f'{get_column_letter(typ_col)}$3:$'
                 f'{get_column_letter(typ_col)}${data_last},"{typ}",'
                 f'{L}$3:{L}${data_last})',
                 font=SKU_FONT, fmt=NUM_FMT_BLANKZERO)

    # ---------------------------------------------------- EDD & Expiry
    ws = wb.create_sheet(TAB_EDD)
    ws.freeze_panes = "D4"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 9
    ws.cell(1, 1, "BEST BEFORE & ESTIMATED DELIVERY DATES  (human-only tab "
                  "— the app never reads this for planning)").font = \
        TITLE_FONT
    ws.cell(1, 4, "Report date:").font = TITLE_FONT
    e1 = ws.cell(1, 5)
    e1.fill, e1.number_format = YELLOW_FILL, DATE_FMT
    exp_col = 4 + 2 * len(regions)
    day_cols = []
    for i, rg in enumerate(regions):
        dc = 4 + 2 * i
        ws.column_dimensions[get_column_letter(dc)].width = 13
        ws.column_dimensions[get_column_letter(dc + 1)].width = 6
        ws.merge_cells(start_row=2, start_column=dc, end_row=2,
                       end_column=dc + 1)
        _set(ws, 2, dc, rg, font=HDR_FONT, fill=HDR_FILL,
             align=Alignment(horizontal="center"))
        ws.cell(2, dc + 1).border = BORDER
        _hdr(ws, 3, dc, "Delivery Date", valign="bottom")
        _hdr(ws, 3, dc + 1, "Days", valign="bottom")
        day_cols.append(dc + 1)
    ws.column_dimensions[get_column_letter(exp_col)].width = 14
    _set(ws, 2, exp_col, "EXPIRY", font=HDR_FONT, fill=GOLD_HDR_FILL,
         align=Alignment(horizontal="center"))
    _hdr(ws, 3, 1, "SKU")
    _hdr(ws, 3, 2, "Product")
    _hdr(ws, 3, 3, "Avg Days")
    _hdr(ws, 3, exp_col, "Best-before date", fill=GOLD_HDR_FILL,
         valign="bottom")
    ws.row_dimensions[3].height = 27.75
    edd_last = 3 + n
    for i, (sku, prod) in enumerate(zip(skus, products)):
        r = 4 + i
        _set(ws, r, 1, sku, font=SKU_FONT)
        _set(ws, r, 2, prod)
        avg_refs = ",".join(f"{get_column_letter(c)}{r}" for c in day_cols)
        _set(ws, r, 3, f'=IFERROR(AVERAGE({avg_refs}),"NA")', fmt="0.0")
        for dc in [c - 1 for c in day_cols]:
            _set(ws, r, dc, None, font=INPUT_FONT, fmt=DATE_FMT)
            DL = get_column_letter(dc)
            _set(ws, r, dc + 1,
                 f'=IF({DL}{r}="","NA",IFERROR({DL}{r}-$E$1,"NA"))', fmt="0")
        _set(ws, r, exp_col, None, font=INPUT_FONT, fmt=DATE_FMT)
    day_rngs = " ".join(f"{get_column_letter(c)}4:{get_column_letter(c)}"
                        f"{edd_last}" for c in [3] + day_cols)
    _cf(ws, day_rngs, "AND(ISNUMBER(C4),C4<2)", CF_GREEN)
    _cf(ws, day_rngs, "AND(ISNUMBER(C4),C4>=2,C4<=3)", CF_AMBER)
    _cf(ws, day_rngs, "AND(ISNUMBER(C4),C4>3)", CF_RED)
    EL = get_column_letter(exp_col)
    _cf(ws, f"{EL}4:{EL}{edd_last}", f'AND({EL}4<>"",{EL}4<$E$1)', CF_RED)
    _cf(ws, f"{EL}4:{EL}{edd_last}", f'AND({EL}4<>"",{EL}4<$E$1+60)',
        CF_AMBER)
    for term, tint in OIL_TINTS:
        _cf(ws, f"B4:B{edd_last}",
            f'NOT(ISERROR(SEARCH("{term}",B4)))', tint)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ================================================================= READER

@dataclass
class StockInputs:
    """Validated content of the completed stock workbook (input tabs only —
    the Overall-stock report tab is NEVER read, spec §7a)."""
    current: pd.DataFrame | None = None      # sku_u, region, qty (incl IXD/YSXA)
    as_on_date: pd.DataFrame | None = None   # sku_u, qty (national, General)
    in_transit: pd.DataFrame | None = None   # shipment_id, region, mode, sku_u,
                                             # qty, cargo_date, reach_date
    at_fc: pd.DataFrame | None = None        # shipment_id, region, sku_u,
                                             # shipped, received, pending
    lead_days: dict = field(default_factory=dict)   # region -> days
    stock_as_of: pd.Timestamp | None = None


def read_stock_workbook(path_or_file, can: Canonical | None = None
                        ) -> tuple[StockInputs, Report]:
    """Parse the completed workbook's INPUT tabs. Every gate is loud:
    blank committed reach date, unknown SKU/destination, bad numbers/dates
    are named errors. Planning must not run unless report.ok."""
    rep = Report()
    out = StockInputs()
    try:
        wb = load_workbook(path_or_file, data_only=True)
    except Exception as e:
        rep.err(f"Stock workbook could not be opened: {e}")
        return out, rep
    for tab in (TAB_REFERENCE, TAB_CURRENT, TAB_INTRANSIT, TAB_ATFC):
        if tab not in wb.sheetnames:
            rep.err(f"Stock workbook is missing the '{tab}' tab "
                    f"(sheets must not be renamed).")
    if rep.errors:
        return out, rep

    known_skus = (set(can.sku_master["sku_u"]) if can is not None and
                  can.sku_master is not None else None)
    known_dests = (set(ordered_regions(can)) | {COL_IXD}
                   if can is not None else None)

    # ---- Reference: lead-days table (runtime home of lead times, §7a)
    ws = wb[TAB_REFERENCE]
    r = 3
    while True:
        rg = ws.cell(r, 1).value
        if rg is None or str(rg).strip() == "":
            break
        days = ws.cell(r, 2).value
        try:
            days = float(days)
            if days < 0 or days > 60:
                raise ValueError
        except (TypeError, ValueError):
            rep.err(f"Reference tab: lead days for '{rg}' is "
                    f"{days!r} — must be a number of days (0–60).")
            r += 1
            continue
        out.lead_days[str(rg).strip()] = days
        r += 1
    if not out.lead_days:
        rep.err("Reference tab: the lead-days table is empty.")
    if known_dests:
        missing = sorted(d for d in known_dests if d not in out.lead_days)
        if missing:
            rep.err(f"Reference tab: no lead days for region(s) {missing}.")

    # ---- Current stock (header-by-name; region set from the file itself)
    ws = wb[TAB_CURRENT]
    out.stock_as_of = _to_date(ws.cell(1, 2).value) \
        if ws.cell(1, 2).value is not None else None
    hdr = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(2, c).value
        if v:
            hdr[str(v).strip()] = c
    if "SKU" not in hdr:
        rep.err(f"'{TAB_CURRENT}': header row 2 has no 'SKU' column.")
        return out, rep
    skip = {"SKU", "Product", "Total", "Current stock (as-on-date)"}
    region_cols = {h: c for h, c in hdr.items() if h not in skip}
    if known_dests:
        unexpected = sorted(set(region_cols) - known_dests - {COL_YSXA})
        if unexpected:
            rep.warn(f"'{TAB_CURRENT}': unrecognized location column(s) "
                     f"{unexpected} — read as-is; check for renamed headers.")
    cur_rows, asod_rows = [], []
    r = 3
    while r <= ws.max_row:
        sku = ws.cell(r, hdr["SKU"]).value
        if sku is None or str(sku).strip() == "":
            break
        sku_u = str(sku).strip().upper()
        if known_skus is not None and sku_u not in known_skus:
            rep.err(f"'{TAB_CURRENT}' row {r}: SKU '{sku}' is not in the "
                    f"master template (STRICT SKU GATE).")
        for rg, c in region_cols.items():
            v = ws.cell(r, c).value
            if v is None or str(v).strip() == "":
                continue
            try:
                q = float(v)
            except (TypeError, ValueError):
                rep.err(f"'{TAB_CURRENT}' row {r}, column '{rg}': "
                        f"{v!r} is not a number.")
                continue
            if q < 0:
                rep.err(f"'{TAB_CURRENT}' row {r}, column '{rg}': "
                        f"negative stock {q}.")
                continue
            if q:
                cur_rows.append({"sku_u": sku_u, "region": rg, "qty": q})
        ac = hdr.get("Current stock (as-on-date)")
        if ac:
            v = ws.cell(r, ac).value
            if v is not None and str(v).strip() != "":
                try:
                    asod_rows.append({"sku_u": sku_u, "qty": float(v)})
                except (TypeError, ValueError):
                    rep.err(f"'{TAB_CURRENT}' row {r}: as-on-date value "
                            f"{v!r} is not a number.")
        r += 1
    out.current = pd.DataFrame(cur_rows)
    out.as_on_date = pd.DataFrame(asod_rows)
    rep.note(f"'{TAB_CURRENT}': {len(cur_rows)} SKU-location balances read"
             + (f", stock-as-of {out.stock_as_of.date()}"
                if out.stock_as_of is not None else ", NO stock-as-of date"))
    if out.stock_as_of is None:
        rep.err(f"'{TAB_CURRENT}': the 'Stock as of' date (cell B1) is "
                f"blank — set it to the Amazon data's date.")

    # ---- In-transit (flat; committed reach date is the gate, §7a)
    ws = wb[TAB_INTRANSIT]
    it_rows = []
    for r in range(3, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 10)]
        sid, dest, mode, sku, _prod, qty, cargo, _sugg, reach = vals
        if all(v is None or str(v).strip() == ""
               for v in (sid, dest, mode, sku, qty, cargo, reach)):
            continue
        rowerr = False
        for name, v in (("Shipment ID", sid), ("Destination", dest),
                        ("Mode", mode), ("SKU", sku), ("Quantity", qty),
                        ("Cargo ready date", cargo)):
            if v is None or str(v).strip() == "":
                rep.err(f"'{TAB_INTRANSIT}' row {r}: {name} is blank — "
                        f"no half-filled rows.")
                rowerr = True
        if reach is None or str(reach).strip() == "":
            rep.err(f"'{TAB_INTRANSIT}' row {r}: 'Actual reach date' is "
                    f"blank — commit a date (the suggestion column is "
                    f"advice, not the input).")
            rowerr = True
        if rowerr:
            continue
        mode = str(mode).strip()
        if mode not in MODES:
            rep.err(f"'{TAB_INTRANSIT}' row {r}: Mode '{mode}' — must be "
                    f"one of {MODES}.")
            continue
        sku_u = str(sku).strip().upper()
        if known_skus is not None and sku_u not in known_skus:
            rep.err(f"'{TAB_INTRANSIT}' row {r}: SKU '{sku}' not in master "
                    f"(STRICT SKU GATE).")
            continue
        dest = str(dest).strip()
        if dest == COL_YSXA:
            rep.err(f"'{TAB_INTRANSIT}' row {r}: YSXA is never a "
                    f"destination.")
            continue
        if known_dests is not None and dest not in known_dests:
            rep.err(f"'{TAB_INTRANSIT}' row {r}: Destination '{dest}' is "
                    f"not a known region.")
            continue
        try:
            q = float(qty)
            if q <= 0:
                raise ValueError
        except (TypeError, ValueError):
            rep.err(f"'{TAB_INTRANSIT}' row {r}: Quantity {qty!r} must be "
                    f"a positive number.")
            continue
        cd, rd = _to_date(cargo), _to_date(reach)
        if pd.isna(cd) or pd.isna(rd):
            rep.err(f"'{TAB_INTRANSIT}' row {r}: unreadable date "
                    f"(cargo={cargo!r}, reach={reach!r}).")
            continue
        if rd < cd:
            rep.warn(f"'{TAB_INTRANSIT}' row {r}: reach date "
                     f"{rd.date()} is before cargo-ready {cd.date()} — "
                     f"check the entry.")
        it_rows.append({"shipment_id": str(sid).strip(), "region": dest,
                        "mode": mode, "sku_u": sku_u, "qty": q,
                        "cargo_date": cd, "reach_date": rd})
    out.in_transit = pd.DataFrame(it_rows)
    rep.note(f"'{TAB_INTRANSIT}': {len(it_rows)} valid shipment lines.")

    # ---- At the FC (numeric shipped/received; available-at-FC = received)
    ws = wb[TAB_ATFC]
    fc_rows = []
    for r in range(3, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 10)]
        sid, dest, sku, _prod, shipped, received, _pend, cargo, recv_d = vals
        if all(v is None or str(v).strip() == ""
               for v in (sid, dest, sku, shipped, received)):
            continue
        rowerr = False
        for name, v in (("Shipment ID", sid), ("Destination", dest),
                        ("SKU", sku), ("Shipped Qty", shipped)):
            if v is None or str(v).strip() == "":
                rep.err(f"'{TAB_ATFC}' row {r}: {name} is blank — no "
                        f"half-filled rows.")
                rowerr = True
        if rowerr:
            continue
        sku_u = str(sku).strip().upper()
        if known_skus is not None and sku_u not in known_skus:
            rep.err(f"'{TAB_ATFC}' row {r}: SKU '{sku}' not in master "
                    f"(STRICT SKU GATE).")
            continue
        dest = str(dest).strip()
        if dest == COL_YSXA:
            rep.err(f"'{TAB_ATFC}' row {r}: YSXA is never a destination.")
            continue
        if known_dests is not None and dest not in known_dests:
            rep.err(f"'{TAB_ATFC}' row {r}: Destination '{dest}' is not a "
                    f"known region.")
            continue
        try:
            sq = float(shipped)
            if sq <= 0:
                raise ValueError
        except (TypeError, ValueError):
            rep.err(f"'{TAB_ATFC}' row {r}: Shipped Qty {shipped!r} must "
                    f"be a positive number.")
            continue
        if received is None or str(received).strip() == "":
            rep.warn(f"'{TAB_ATFC}' row {r}: Received Qty blank — treated "
                     f"as 0 received (all {sq:.0f} still pending).")
            rq = 0.0
        else:
            try:
                rq = float(received)
                if rq < 0:
                    raise ValueError
            except (TypeError, ValueError):
                rep.err(f"'{TAB_ATFC}' row {r}: Received Qty {received!r} "
                        f"must be a number (>= 0) — write quantities, not "
                        f"words like 'fully received'.")
                continue
        if rq > sq:
            rep.warn(f"'{TAB_ATFC}' row {r}: Received {rq:.0f} exceeds "
                     f"Shipped {sq:.0f} — check the entry.")
        fc_rows.append({"shipment_id": str(sid).strip(), "region": dest,
                        "sku_u": sku_u, "shipped": sq, "received": rq,
                        "pending": sq - rq})
    out.at_fc = pd.DataFrame(fc_rows)
    rep.note(f"'{TAB_ATFC}': {len(fc_rows)} valid receiving lines.")
    return out, rep
