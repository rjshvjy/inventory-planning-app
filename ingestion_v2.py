"""
ingestion_v2.py — Extract & Validate layer for the Inventory Planning App v2.

Reads repo reference files (configurations.xlsx, inventory_plan_template.xlsx,
fc_registration.pdf/.csv) and the per-run Amazon exports (sales CSV, General
stock CSV, FC-wise ledger CSV), applies every validation gate from the
Redesign Spec (v2.7), and returns canonical data structures + a validation
report. Business logic never touches raw files — it consumes this module's
output only.

Design contract (spec refs):
  §4  master template = SKU single source of truth; STRICT unknown-SKU halt
  §5  FC→region from registration file (content-based parse, fail loud)
  §5a Region_Splits overlay (destination-only; stock pooling per split region)
  §6  sales: PII dropped on read; daily = qty ÷ actual days in window
  §6a data-date anchor + cross-file freshness spread
  §7  stock: ledger=available per FC (SELLABLE only); General=national states
  §7b content-based column detection, loud stops, never guess
"""

from __future__ import annotations
import io
import re
from dataclasses import dataclass, field
from datetime import datetime, date

import pandas as pd
from openpyxl import load_workbook

# ---------------------------------------------------------------- constants

# Display names for supply regions, keyed by Amazon 2-letter state code.
# Unknown codes fall back to the code itself (warning, not a stop) so a newly
# registered state still auto-creates a column (spec v2.6 change 4).
REGION_NAMES = {
    "DL": "Delhi", "HR": "Haryana", "MH": "Bombay", "KA": "Bangalore",
    "TG": "Hyderabad", "TN": "Chennai", "WB": "Calcutta", "UP": "Lucknow",
    "GJ": "Ahmedabad", "RJ": "Jaipur", "PB": "Ludhiana", "MP": "Bhopal",
    "AP": "Vijayawada", "KL": "Kochi", "BR": "Patna", "OR": "Bhubaneswar",
    "AS": "Guwahati",
}

INDIAN_STATES = {  # for demand-state normalization & content detection
    "ANDHRA PRADESH", "ARUNACHAL PRADESH", "ASSAM", "BIHAR", "CHHATTISGARH",
    "GOA", "GUJARAT", "HARYANA", "HIMACHAL PRADESH", "JHARKHAND", "KARNATAKA",
    "KERALA", "MADHYA PRADESH", "MAHARASHTRA", "MANIPUR", "MEGHALAYA",
    "MIZORAM", "NAGALAND", "ODISHA", "PUNJAB", "RAJASTHAN", "SIKKIM",
    "TAMIL NADU", "TELANGANA", "TRIPURA", "UTTAR PRADESH", "UTTARAKHAND",
    "WEST BENGAL", "DELHI", "JAMMU AND KASHMIR", "JAMMU & KASHMIR", "LADAKH",
    "PUDUCHERRY", "CHANDIGARH", "ANDAMAN AND NICOBAR ISLANDS", "LAKSHADWEEP",
    "DADRA AND NAGAR HAVELI", "DAMAN AND DIU",
    "DADRA AND NAGAR HAVELI AND DAMAN AND DIU",
}
STATE_VARIANTS = {  # spelling drift seen in real Amazon exports -> canonical
    "PONDICHERRY": "PUDUCHERRY", "ORISSA": "ODISHA",
    "JAMMU & KASHMIR": "JAMMU AND KASHMIR", "NEW DELHI": "DELHI",
    "DADRA & NAGAR HAVELI": "DADRA AND NAGAR HAVELI",
    "TAMILNADU": "TAMIL NADU", "TAMIL NADU.": "TAMIL NADU",
    "CHATTISGARH": "CHHATTISGARH", "TELENGANA": "TELANGANA",
    "TELANGANA STATE": "TELANGANA", "KERAL STATE": "KERALA", "KERELA": "KERALA",
    "U.P": "UTTAR PRADESH", "U.P.": "UTTAR PRADESH", "UP": "UTTAR PRADESH",
    "M.P": "MADHYA PRADESH", "MP": "MADHYA PRADESH",
    # customers sometimes type the 2-letter code as the state
    "AP": "ANDHRA PRADESH", "HR": "HARYANA", "MH": "MAHARASHTRA",
    "TG": "TELANGANA", "TN": "TAMIL NADU", "KA": "KARNATAKA", "DL": "DELHI",
    "WB": "WEST BENGAL", "KL": "KERALA", "GJ": "GUJARAT", "RJ": "RAJASTHAN",
}

STATE_NAME_TO_CODE = {  # canonical spelled state -> 2-letter code (fact of the
    # world, same family as INDIAN_STATES; used by Demand_Map validation §5b and
    # by calculation's demand-state -> region resolver)
    "ANDHRA PRADESH": "AP", "ARUNACHAL PRADESH": "AR", "ASSAM": "AS",
    "BIHAR": "BR", "CHHATTISGARH": "CG", "GOA": "GA", "GUJARAT": "GJ",
    "HARYANA": "HR", "HIMACHAL PRADESH": "HP", "JHARKHAND": "JH",
    "KARNATAKA": "KA", "KERALA": "KL", "MADHYA PRADESH": "MP",
    "MAHARASHTRA": "MH", "MANIPUR": "MN", "MEGHALAYA": "ML", "MIZORAM": "MZ",
    "NAGALAND": "NL", "ODISHA": "OR", "PUNJAB": "PB", "RAJASTHAN": "RJ",
    "SIKKIM": "SK", "TAMIL NADU": "TN", "TELANGANA": "TG", "TRIPURA": "TR",
    "UTTAR PRADESH": "UP", "UTTARAKHAND": "UK", "WEST BENGAL": "WB",
    "DELHI": "DL", "JAMMU AND KASHMIR": "JK", "LADAKH": "LA",
    "PUDUCHERRY": "PY", "CHANDIGARH": "CH",
    "ANDAMAN AND NICOBAR ISLANDS": "AN", "LAKSHADWEEP": "LD",
    "DADRA AND NAGAR HAVELI": "DN", "DAMAN AND DIU": "DD",
    "DADRA AND NAGAR HAVELI AND DAMAN AND DIU": "DH",
}

FC_CODE_RE = re.compile(r"^[A-Z]{3,4}\d?$|^[A-Z]{2}\d[A-Z0-9]$|^[A-Z]{4}$")
CONFIG_KEYS = {  # Config tab: required parameter -> (type, sane-range check)
    "days_cover_ceiling": (float, lambda v: 1 <= v <= 365),
    "min_sales_window_days": (float, lambda v: 1 <= v <= 180),
    "freshness_tolerance_days": (float, lambda v: 0 <= v <= 30),
    "min_stockout_deficit_units": (float, lambda v: 0 <= v <= 1000),
    "sales_recency_prompt_days": (float, lambda v: 0 < v <= 365),  # v2.9 §6a
}

# ---------------------------------------------------------------- reporting


@dataclass
class Report:
    """Validation report. Any error blocks the run (fail loud, fail specific)."""
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    info: list[str] = field(default_factory=list)
    # v2.9 interactive gates — NOT errors; app must resolve before calculating
    needs_fc_review: dict = field(default_factory=dict)   # fc -> sellable units
    needs_recency_ack: dict = field(default_factory=dict) # {sales,stock,gap_days}

    def err(self, msg): self.errors.append(msg)
    def warn(self, msg): self.warnings.append(msg)
    def note(self, msg): self.info.append(msg)
    @property
    def ok(self): return not self.errors


@dataclass
class Canonical:
    """Validated canonical data the calculation layer consumes."""
    config: dict = field(default_factory=dict)
    fc_types: dict = field(default_factory=dict)          # FC code -> IXD|IGNORE
    region_splits: list = field(default_factory=list)      # rows of split table
    demand_map: dict = field(default_factory=dict)          # non-FC demand state -> serving state code (§5b)
    fc_resolutions: dict = field(default_factory=dict)      # v2.9 §5c: per-run user FC mappings as applied
    fc_region: dict = field(default_factory=dict)          # FC code -> region label
    regions: list = field(default_factory=list)            # ordered region labels
    sku_master: pd.DataFrame | None = None                 # per-SKU master data
    sales_daily: pd.DataFrame | None = None                # sku x state daily sales
    sales_window_days: float | None = None
    national: pd.DataFrame | None = None                   # per-SKU national states
    available: pd.DataFrame | None = None                  # sku x region SELLABLE
    effective_dates: dict = field(default_factory=dict)    # source -> date


# ------------------------------------------------------------ small helpers

def _find_col(df: pd.DataFrame, names: list[str], content=None, label="",
              rep: Report | None = None) -> str | None:
    """Header-match first (case/space-insensitive), content-match second,
    loud None third (spec §7b layered detection)."""
    norm = {re.sub(r"\s+", " ", str(c)).strip().lower(): c for c in df.columns}
    for n in names:
        if n.lower() in norm:
            return norm[n.lower()]
    if content is not None:
        best, score = None, 0.0
        sample = df.head(200)
        for c in df.columns:
            vals = sample[c].dropna().astype(str)
            if len(vals) == 0:
                continue
            s = sum(bool(content(v)) for v in vals) / len(vals)
            if s > score:
                best, score = c, s
        if score >= 0.6:
            if rep:
                rep.warn(f"{label}: header not found by name; matched column "
                         f"'{best}' by content ({score:.0%} of values fit).")
            return best
    return None


def _read_csv(src) -> pd.DataFrame:
    """Read a CSV/TSV upload or path; sniff tab vs comma."""
    if hasattr(src, "read"):
        raw = src.read()
        if isinstance(raw, bytes):
            raw = raw.decode("utf-8", errors="replace")
    else:
        raw = open(src, encoding="utf-8", errors="replace").read()
    delim = "\t" if raw[:4000].count("\t") > raw[:4000].count(",") else ","
    return pd.read_csv(io.StringIO(raw), delimiter=delim)


def _to_date(v):
    if isinstance(v, (datetime, date)):
        ts = pd.Timestamp(v)
    else:
        ts = pd.to_datetime(str(v), errors="coerce", dayfirst=False)
    if ts is pd.NaT or ts is None:
        return pd.NaT
    if getattr(ts, "tzinfo", None) is not None:
        ts = ts.tz_localize(None)
    return ts.normalize()


# ---------------------------------------------------- reference file loaders

def load_config(path_or_file, rep: Report) -> tuple[dict, dict, list, dict]:
    """configurations.xlsx -> (config numbers, fc_types, region_splits,
    demand_map). Demand_Map (§5b) is structurally validated here; the
    FC-dependent checks (serving code has FCs, no FC-state rows) run in
    run_ingestion after the registration loads."""
    cfg, fct, splits, dmap = {}, {}, [], {}
    try:
        wb = load_workbook(path_or_file, data_only=True)
    except Exception as e:
        rep.err(f"configurations.xlsx could not be opened: {e}")
        return cfg, fct, splits, dmap

    for tab in ("Config", "FC_Types", "Region_Splits", "Demand_Map"):
        if tab not in wb.sheetnames:
            rep.err(f"configurations.xlsx is missing the '{tab}' tab.")
    if rep.errors:
        return cfg, fct, splits, dmap

    ws = wb["Config"]
    for r in range(2, ws.max_row + 1):
        k, v = ws.cell(r, 1).value, ws.cell(r, 2).value
        if k is None:
            continue
        cfg[str(k).strip()] = v
    for key, (typ, check) in CONFIG_KEYS.items():
        if key not in cfg or cfg[key] is None or str(cfg[key]).strip() == "":
            rep.err(f"Config tab: '{key}' is missing or blank.")
            continue
        try:
            cfg[key] = typ(cfg[key])
        except (TypeError, ValueError):
            rep.err(f"Config tab: '{key}' = {cfg[key]!r} is not a number.")
            continue
        if not check(cfg[key]):
            rep.err(f"Config tab: '{key}' = {cfg[key]} is outside its sane range.")
    for k in cfg:
        if k not in CONFIG_KEYS:
            rep.warn(f"Config tab: unknown parameter '{k}' (ignored).")

    ws = wb["FC_Types"]
    for r in range(2, ws.max_row + 1):
        code, typ = ws.cell(r, 1).value, ws.cell(r, 2).value
        if code is None:
            continue
        code, typ = str(code).strip().upper(), str(typ or "").strip().upper()
        if typ not in ("IXD", "IGNORE"):
            rep.err(f"FC_Types tab: FC '{code}' has type '{typ}' — must be "
                    f"exactly IXD or IGNORE (typo stops the run, never a guess).")
        fct[code] = typ

    ws = wb["Region_Splits"]
    for r in range(2, ws.max_row + 1):
        st = ws.cell(r, 1).value
        if st is None:
            continue
        splits.append({
            "state": str(st).strip().upper(),
            "label": str(ws.cell(r, 2).value or "").strip(),
            "prefixes": [p.strip().upper() for p in
                         str(ws.cell(r, 3).value or "").split(",") if p.strip()],
            "default": str(ws.cell(r, 4).value or "").strip().upper() == "YES",
        })
    for st in {s["state"] for s in splits}:
        rows = [s for s in splits if s["state"] == st]
        ndef = sum(s["default"] for s in rows)
        if ndef != 1:
            rep.err(f"Region_Splits: state {st} has {ndef} DEFAULT rows — "
                    f"exactly one region must be marked YES.")
        for s in rows:
            if not s["label"] or not s["prefixes"]:
                rep.err(f"Region_Splits: state {st} row '{s['label']}' is "
                        f"missing a label or FC prefixes.")

    ws = wb["Demand_Map"]
    for r in range(2, ws.max_row + 1):
        st = ws.cell(r, 1).value
        if st is None or str(st).strip() == "":
            continue
        state = str(st).strip().upper()
        state = STATE_VARIANTS.get(state, state)        # tolerate spelling drift
        code = str(ws.cell(r, 2).value or "").strip().upper()
        if state not in INDIAN_STATES:
            rep.err(f"Demand_Map: '{st}' is not a recognized Indian state/UT "
                    f"name (row {r}).")
            continue
        if state in dmap:
            rep.err(f"Demand_Map: duplicate row for '{state}' (row {r}) — "
                    f"each demand state may appear once.")
            continue
        if not code or len(code) != 2 or not code.isalpha():
            rep.err(f"Demand_Map: '{state}' has serving code {code!r} — must "
                    f"be a 2-letter state code (row {r}).")
            continue
        dmap[state] = code
    return cfg, fct, splits, dmap


def load_master(path_or_file, rep: Report) -> pd.DataFrame | None:
    """Master template 'Appointment plan' -> SKU master dataframe.
    Parsing contract (§4): from first SKU row down, stop at first blank."""
    try:
        wb = load_workbook(path_or_file, data_only=True)
    except Exception as e:
        rep.err(f"Master template could not be opened: {e}")
        return None
    if "Appointment plan" not in wb.sheetnames:
        rep.err("Master template: sheet 'Appointment plan' not found "
                "(do not rename it).")
        return None
    ws = wb["Appointment plan"]
    hdr = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(2, c).value
        if v:
            hdr[str(v).strip()] = c
    need = ["SKU", "ASIN", "Units/Box", "Bottles/Unit", "Litres/Unit",
            "Kg/Unit", "Oil Type", "Bottle Type"]
    missing = [n for n in need if n not in hdr]
    if missing:
        rep.err(f"Master template header row 2 is missing: {missing} "
                f"(headers must not be renamed).")
        return None

    rows, r = [], 3
    started = False
    while r <= ws.max_row + 1:
        sku = ws.cell(r, hdr["SKU"]).value
        if sku is None or str(sku).strip() == "":
            if started:
                break              # contract: stop at first blank after start
            r += 1
            continue
        started = True
        rows.append({
            "sku": str(sku).strip(),
            "asin": str(ws.cell(r, hdr["ASIN"]).value or "").strip(),
            "alias": str(ws.cell(r, hdr.get("Alias / Display name",
                                            hdr["SKU"])).value or "").strip(),
            "item": str(ws.cell(r, hdr.get("ITEM", hdr["SKU"])).value or "").strip(),
            "units_per_box": ws.cell(r, hdr["Units/Box"]).value,
            "bottles_per_unit": ws.cell(r, hdr["Bottles/Unit"]).value,
            "litres_per_unit": ws.cell(r, hdr["Litres/Unit"]).value,
            "kg_per_unit": ws.cell(r, hdr["Kg/Unit"]).value,
            "oil_type": str(ws.cell(r, hdr["Oil Type"]).value or "").strip(),
            "bottle_type": str(ws.cell(r, hdr["Bottle Type"]).value or "").strip(),
            "row": r,
        })
        r += 1
    df = pd.DataFrame(rows)
    if df.empty:
        rep.err("Master template: no SKU rows found under the header.")
        return None
    dups = df[df["sku"].str.upper().duplicated()]["sku"].tolist()
    if dups:
        rep.err(f"Master template: duplicate SKUs {dups}.")
    bad = df[~pd.to_numeric(df["units_per_box"], errors="coerce").gt(0)]
    if len(bad):
        rep.err(f"Master template: Units/Box missing or not positive for "
                f"{bad['sku'].tolist()}.")
    for c in ("units_per_box", "bottles_per_unit", "litres_per_unit",
              "kg_per_unit"):
        df[c] = pd.to_numeric(df[c], errors="coerce")
    df["sku_u"] = df["sku"].str.upper()
    rep.note(f"Master: {len(df)} SKUs, {df['asin'].nunique()} ASINs "
             f"({len(df) - df['asin'].nunique()} Temp/shared-ASIN rows).")
    return df


def load_fc_registration(path_or_file, fc_types: dict, splits: list,
                         rep: Report) -> tuple[dict, list]:
    """FC registration (pdf or csv) -> fc_region map + ordered region labels.
    Content-based: find (2-letter state code, FC code) pairs; fail loud."""
    pairs = []
    name = getattr(path_or_file, "name", str(path_or_file)).lower()
    try:
        if name.endswith(".csv"):
            df = _read_csv(path_or_file)
            sc = _find_col(df, ["state"], lambda v: len(v.strip()) == 2
                           and v.strip().isalpha(), "FC file state col", rep)
            fc = _find_col(df, ["fulfillment center", "fc"],
                           lambda v: bool(FC_CODE_RE.match(v.strip().upper())),
                           "FC file FC col", rep)
            if not sc or not fc:
                rep.err("FC registration CSV: could not locate the State and "
                        "FC columns by name or content.")
                return {}, []
            for _, r in df.iterrows():
                pairs.append((str(r[sc]).strip().upper(),
                              str(r[fc]).strip().upper()))
        else:  # pdf
            import pdfplumber
            with pdfplumber.open(path_or_file) as pdf:
                text = "\n".join(p.extract_text() or "" for p in pdf.pages)
            for m in re.finditer(r"\b([A-Z]{2})\s+([A-Z]{2,4}\d?[A-Z0-9]?)\b",
                                 text):
                st, fc = m.group(1), m.group(2)
                if FC_CODE_RE.match(fc) and st not in ("EN",):
                    pairs.append((st, fc))
    except Exception as e:
        rep.err(f"FC registration file could not be parsed: {e}")
        return {}, []

    seen, clean = set(), []
    for st, fc in pairs:
        if fc not in seen and len(fc) >= 3:
            seen.add(fc)
            clean.append((st, fc))
    if len(clean) < 5:
        rep.err(f"FC registration: only {len(clean)} FC rows recognized — "
                f"file format may have changed. First pairs seen: {clean[:5]}. "
                f"Check the export or supply a 2-column CSV (State, FC).")
        return {}, []

    fc_region, regions = {}, []

    def add_region(lbl):
        if lbl not in regions:
            regions.append(lbl)

    split_states = {s["state"] for s in splits}
    for st, fc in clean:
        if fc_types.get(fc) == "IGNORE":
            fc_region[fc] = "YSXA"
            continue
        if fc_types.get(fc) == "IXD":
            fc_region[fc] = "BLR4 IXD"
            continue
        if st in split_states:
            hit = [s for s in splits if s["state"] == st
                   and any(fc.startswith(p) for p in s["prefixes"])]
            if len(hit) == 1:
                fc_region[fc] = f"{hit[0]['label']} ({st})"
            elif len(hit) > 1:
                rep.err(f"Region_Splits: FC {fc} matches more than one region "
                        f"of state {st} ({[h['label'] for h in hit]}).")
            else:
                dft = next(s for s in splits if s["state"] == st and s["default"])
                fc_region[fc] = f"{dft['label']} ({st})"
        else:
            cap = REGION_NAMES.get(st)
            if cap is None:
                cap = st
                rep.warn(f"FC registration: state code '{st}' has no display "
                         f"name on file — labelled '{st} ({st})'. Add it to "
                         f"REGION_NAMES when convenient.")
            fc_region[fc] = f"{cap} ({st})"

    # ordered region list: registration order, splits expanded, specials last
    for st, fc in clean:
        lbl = fc_region[fc]
        if lbl not in ("YSXA", "BLR4 IXD"):
            add_region(lbl)
    rep.note(f"FC registration: {len(clean)} FCs across "
             f"{len(regions)} regions "
             f"(+ IXD/{sum(1 for v in fc_region.values() if v=='BLR4 IXD')}"
             f", ignore/{sum(1 for v in fc_region.values() if v=='YSXA')}).")
    return fc_region, regions


# --------------------------------------------------------- Amazon CSV loaders

def load_sales(src, master: pd.DataFrame, cfg: dict, rep: Report,
               window_days_override: float | None = None):
    """Sales export -> (daily sales sku x state, window_days, max_date).
    PII is dropped immediately: only SKU, qty, state, date survive the read."""
    df = _read_csv(src)
    sku_c = _find_col(df, ["merchant sku", "sku", "seller-sku", "seller sku"],
                      None, "sales SKU", rep)
    qty_c = _find_col(df, ["shipped quantity", "quantity-shipped", "quantity",
                           "qty"], None, "sales qty", rep)
    st_c = _find_col(df, ["shipping state", "ship-state", "state"],
                     lambda v: v.strip().upper() in INDIAN_STATES
                     or v.strip().upper() in STATE_VARIANTS,
                     "sales state", rep)
    if not (sku_c and qty_c and st_c):
        rep.err(f"Sales file: could not locate required columns "
                f"(found sku={sku_c}, qty={qty_c}, state={st_c}). "
                f"Columns present: {list(df.columns)[:12]}")
        return None, None, None
    dt_c = _find_col(df, ["shipment date", "shipments date", "purchase date",
                          "payments date"], None, "sales date", None)
    stat_c = _find_col(df, ["shipment status", "order status", "status"],
                       None, "sales status", None)

    keep = [sku_c, qty_c, st_c] + ([dt_c] if dt_c else []) \
        + ([stat_c] if stat_c else [])
    df = df[keep].copy()                                   # PII dropped here
    if stat_c:
        before = len(df)
        df = df[~df[stat_c].astype(str).str.upper()
                .str.contains("CANCEL|RETURN", na=False)]
        if before - len(df):
            rep.note(f"Sales: filtered {before - len(df)} cancelled/returned "
                     f"lines.")

    df["state"] = (df[st_c].astype(str).str.strip().str.upper()
                   .replace(STATE_VARIANTS))
    unknown = sorted(set(df.loc[~df["state"].isin(INDIAN_STATES), "state"])
                     - {"NAN", ""})
    if unknown:
        rep.warn(f"Sales: unrecognized state spellings kept as-is: {unknown}.")
    df["sku_u"] = df[sku_c].astype(str).str.strip().str.upper()
    df["qty"] = pd.to_numeric(df[qty_c], errors="coerce").fillna(0)

    # strict SKU gate (§4)
    unknown_skus = sorted(set(df["sku_u"]) - set(master["sku_u"]))
    if unknown_skus:
        rep.err("STRICT SKU GATE — sales file contains SKU(s) not in the "
                f"master template: {unknown_skus}. Add them to the master "
                f"(one row each, with ASIN and box data), commit, and rerun. "
                f"No partial plan is produced.")

    # window
    max_date = None
    if dt_c:
        dts = df[dt_c].map(_to_date).dropna()
        if len(dts):
            window = (dts.max() - dts.min()).days + 1
            max_date = dts.max()
            rep.note(f"Sales window inferred from dates: "
                     f"{dts.min().date()} → {dts.max().date()} "
                     f"({window} days).")
        else:
            window = None
    else:
        window = None
    if window is None:
        if window_days_override:
            window = float(window_days_override)
            rep.warn(f"Sales file has no usable date column — using the "
                     f"manually entered window of {window:.0f} days.")
        else:
            rep.err("Sales file has no date column and no window was entered. "
                    "Either export the report with dates or enter the number "
                    "of days it covers.")
            return None, None, None
    if window < cfg.get("min_sales_window_days", 30):
        rep.warn(f"Sales window is {window:.0f} days — below the "
                 f"{cfg['min_sales_window_days']:.0f}-day minimum; averages "
                 f"for slow movers will be noisy.")

    g = (df.groupby(["sku_u", "state"])["qty"].sum().reset_index())
    g["daily"] = g["qty"] / window
    return g, window, max_date


def load_general_stock(src, master: pd.DataFrame, rep: Report):
    """Manage FBA Inventory CSV -> per-SKU national inventory states."""
    df = _read_csv(src)
    cols = {
        "sku": ["sku", "seller-sku", "merchant sku"],
        "fulfillable": ["afn-fulfillable-quantity"],
        "reserved": ["afn-reserved-quantity"],
        "inbound_shipped": ["afn-inbound-shipped-quantity"],
        "inbound_receiving": ["afn-inbound-receiving-quantity"],
        "inbound_working": ["afn-inbound-working-quantity"],
        "asin": ["asin"],
    }
    found = {k: _find_col(df, v, None, f"general {k}", rep)
             for k, v in cols.items()}
    missing = [k for k in ("sku", "fulfillable") if not found[k]]
    if missing:
        rep.err(f"General stock file: required column(s) not found: {missing}."
                f" Columns present: {list(df.columns)[:12]}")
        return None
    out = pd.DataFrame({"sku_u": df[found["sku"]].astype(str)
                        .str.strip().str.upper()})
    for k in ("fulfillable", "reserved", "inbound_shipped",
              "inbound_receiving", "inbound_working"):
        out[k] = (pd.to_numeric(df[found[k]], errors="coerce").fillna(0)
                  if found[k] else 0)
    unknown = sorted(set(out["sku_u"]) - set(master["sku_u"]))
    if unknown:
        rep.err(f"STRICT SKU GATE — General stock file contains SKU(s) not in "
                f"the master template: {unknown}. Add to master, commit, rerun.")
    out = out.groupby("sku_u", as_index=False).sum(numeric_only=True)
    rep.note(f"General stock: {len(out)} SKUs; national fulfillable "
             f"{int(out['fulfillable'].sum())}, reserved "
             f"{int(out['reserved'].sum())}, inbound "
             f"{int(out['inbound_shipped'].sum())} shipped / "
             f"{int(out['inbound_receiving'].sum())} receiving.")
    return out


def load_ledger(src, master: pd.DataFrame, fc_region: dict, rep: Report):
    """FC-wise Inventory Ledger CSV -> (available sku x region, ledger_date).
    SELLABLE only; unknown FC holding stock = loud stop (§5 known-answer)."""
    df = _read_csv(src)
    sku_c = _find_col(df, ["msku", "sku"], None, "ledger sku", rep)
    loc_c = _find_col(df, ["location", "fulfillment center", "fc"],
                      lambda v: bool(FC_CODE_RE.match(v.strip().upper())),
                      "ledger location", rep)
    bal_c = _find_col(df, ["ending warehouse balance"], None,
                      "ledger balance", rep)
    dis_c = _find_col(df, ["disposition"], None, "ledger disposition", rep)
    dat_c = _find_col(df, ["date"], None, "ledger date", None)
    if not (sku_c and loc_c and bal_c and dis_c):
        rep.err(f"Ledger file: required columns not found "
                f"(sku={sku_c}, location={loc_c}, balance={bal_c}, "
                f"disposition={dis_c}). Columns: {list(df.columns)[:12]}")
        return None, None
    ledger_date = None
    if dat_c:
        d = df[dat_c].map(_to_date).dropna()
        if len(d):
            ledger_date = d.max()

    df = df[df[dis_c].astype(str).str.upper() == "SELLABLE"].copy()
    df["fc"] = df[loc_c].astype(str).str.strip().str.upper()
    df["sku_u"] = df[sku_c].astype(str).str.strip().str.upper()
    df["bal"] = pd.to_numeric(df[bal_c], errors="coerce").fillna(0)

    unknown_fc = sorted(set(df.loc[df["bal"] > 0, "fc"]) - set(fc_region))
    for fc in unknown_fc:                       # v2.9 §5c: interactive gate,
        units = int(df.loc[(df["fc"] == fc), "bal"].sum())   # not a hard stop
        rep.needs_fc_review[fc] = units
        rep.warn(f"FC-review needed: '{fc}' holds {units} sellable units but "
                 f"is not in the registration file — the app will ask you to "
                 f"map / IXD / ignore / abort (per-run; permanent fix = "
                 f"commit a refreshed registration export).")
    unknown_skus = sorted(set(df.loc[df["bal"] > 0, "sku_u"])
                          - set(master["sku_u"]))
    if unknown_skus:
        rep.err(f"STRICT SKU GATE — ledger contains SKU(s) not in the master "
                f"template: {unknown_skus}. Add to master, commit, rerun.")

    df["region"] = df["fc"].map(fc_region)
    # YSXA is KEPT here for display (stock workbook shows what's parked
    # there); it is never planning supply — calculation iterates
    # canonical.regions, which excludes YSXA and BLR4 IXD.
    avail = (df[df["region"].notna()]
             .groupby(["sku_u", "region"])["bal"].sum().reset_index())
    rep.note(f"Ledger: SELLABLE balances across "
             f"{df['fc'].nunique()} FCs"
             + (f", data date {ledger_date.date()}" if ledger_date is not None
                else ", no date column found"))
    return avail, ledger_date


# ------------------------------------------------------------- orchestration

def run_ingestion(sales_file, general_file, ledger_file,
                  config_path="reference/configurations.xlsx",
                  master_path="reference/inventory_plan_template.xlsx",
                  fcreg_path="reference/fc_registration.pdf",
                  window_days_override: float | None = None,
                  today: pd.Timestamp | None = None,
                  fc_resolutions: dict | None = None,
                  recency_acknowledged: bool = False
                  ) -> tuple[Canonical, Report]:
    """Full Extract+Validate pass. Returns (canonical data, report).
    Calculation must not run unless report.ok is True AND the v2.9 gate
    fields (rep.needs_fc_review, rep.needs_recency_ack) are empty/handled.

    v2.9 (§5c/§6a): fc_resolutions = per-run user answers from the FC-review
    gate, e.g. {"MAA4": {"action": "map", "region": "Chennai (TN)",
    "fulfillable": True}} (actions: map / ixd / ignore). recency_acknowledged
    = True suppresses the sales-recency prompt after the user confirms."""
    rep = Report()
    can = Canonical()
    today = today or pd.Timestamp.now().normalize()

    can.config, can.fc_types, can.region_splits, can.demand_map = load_config(
        config_path, rep)
    can.sku_master = load_master(master_path, rep)
    if rep.errors:                       # reference layer broken: stop early
        return can, rep
    can.fc_region, can.regions = load_fc_registration(
        fcreg_path, can.fc_types, can.region_splits, rep)
    if rep.errors:
        return can, rep

    # Demand_Map FC-dependent validation (§5b) — needs the registration
    fc_state_codes = {m.group(1) for lbl in can.fc_region.values()
                      if lbl not in ("YSXA", "BLR4 IXD")
                      and (m := re.search(r"\((\w{2})\)$", lbl))}
    code_to_name = {v: k for k, v in STATE_NAME_TO_CODE.items()}
    for state, code in list(can.demand_map.items()):
        if code not in fc_state_codes:
            rep.err(f"Demand_Map: '{state}' -> {code}, but {code} has no "
                    f"registered FCs — demand cannot be served from a state "
                    f"with no FC.")
        own = STATE_NAME_TO_CODE.get(state)
        if own in fc_state_codes:
            rep.warn(f"Demand_Map: '{state}' now has its own registered FCs — "
                     f"its row is ignored (maps to itself automatically); "
                     f"delete it from configurations.xlsx.")
            del can.demand_map[state]
    if can.demand_map:
        rep.note(f"Demand_Map: {len(can.demand_map)} non-FC demand states "
                 f"mapped to serving regions "
                 f"({sorted(set(can.demand_map.values()))}).")

    # v2.9 §5c — apply per-run FC resolutions BEFORE the ledger is read, so
    # resolved FCs map cleanly and only unresolved ones raise the gate.
    for fc, res in (fc_resolutions or {}).items():
        fc, act = fc.strip().upper(), str(res.get("action", "")).lower()
        if act == "map":
            region = res.get("region")
            if region not in can.regions:
                rep.err(f"FC resolution: '{fc}' mapped to unknown region "
                        f"{region!r} — pick from {can.regions}.")
                continue
            if res.get("fulfillable", True):
                can.fc_region[fc] = region
                rep.note(f"FC-review: {fc} mapped to {region} (fulfillable) "
                         f"by user for this run — flagged in output.")
            else:
                can.fc_region[fc] = f"{fc} EXCLUDED({region})"
                rep.warn(f"FC-review: {fc} mapped to {region} but marked NOT "
                         f"fulfillable — its stock is shown, excluded from "
                         f"planning supply this run.")
        elif act == "ixd":
            can.fc_region[fc] = "BLR4 IXD"
            rep.warn(f"FC-review: {fc} treated as IXD by user — its stock "
                     f"joins the cross-dock pool this run.")
        elif act == "ignore":
            can.fc_region[fc] = f"{fc} IGNORED"
            rep.warn(f"FC-review: {fc} IGNORED by user this run — stock "
                     f"shown, never counted.")
        else:
            rep.err(f"FC resolution for '{fc}': unknown action {act!r} "
                    f"(map / ixd / ignore).")
        can.fc_resolutions[fc] = dict(res)

    can.sales_daily, can.sales_window_days, sales_max = load_sales(
        sales_file, can.sku_master, can.config, rep, window_days_override)
    can.national = load_general_stock(general_file, can.sku_master, rep)
    can.available, ledger_date = load_ledger(
        ledger_file, can.sku_master, can.fc_region, rep)

    # date-staleness & cross-file freshness (§6a)
    if ledger_date is not None:
        can.effective_dates["ledger"] = ledger_date
        lag = (today - ledger_date).days
        rep.note(f"Planning anchor = ledger data date {ledger_date.date()} "
                 f"({lag} days behind today — Amazon runs ~2 days behind).")
    if sales_max is not None:
        can.effective_dates["sales"] = sales_max
    # v2.9 §6a freshness SPLIT: sales-vs-stock is an acknowledge gate, not
    # an error. (Stock-vs-stock pairs are checked where both dates exist —
    # here only the ledger carries a stock date; the workbook's stock-as-of
    # is compared against it by the workbook reader.)
    if "ledger" in can.effective_dates and "sales" in can.effective_dates:
        gap = (can.effective_dates["ledger"]
               - can.effective_dates["sales"]).days
        prompt_at = can.config.get("sales_recency_prompt_days", 60)
        if gap > prompt_at and not recency_acknowledged:
            rep.needs_recency_ack = {
                "sales_end": str(can.effective_dates["sales"].date()),
                "stock_anchor": str(can.effective_dates["ledger"].date()),
                "gap_days": gap, "threshold": prompt_at}
            rep.warn(f"Sales-recency: window ends "
                     f"{can.effective_dates['sales'].date()}, stock anchor "
                     f"{can.effective_dates['ledger'].date()} ({gap} days) — "
                     f"exceeds {prompt_at:.0f}d; the app will ask you to "
                     f"confirm this velocity as current before planning.")
        elif gap > prompt_at and recency_acknowledged:
            rep.note(f"Sales-recency acknowledged by user: {gap}-day-old "
                     f"window applied as current velocity — flagged in "
                     f"output.")
    return can, rep
