"""
golden_harness_v2.py — headless golden-test harness for calculation_v2.

WHY THIS EXISTS (written at ~75% context of window 3, 8 Jul 2026): getting the
July-7 sample data through the pipeline needs three fixtures whose exact
construction lived only in that window's context. This script rebuilds them
DETERMINISTICALLY from the raw uploads, runs the full pipeline headlessly, and
dumps a golden-comparison workbook. Next window: fill GOLDEN below with
known-good numbers, run, compare. No Streamlit, no fixture archaeology.

USAGE
    python3 golden_harness_v2.py [uploads_dir] [out_dir]
    (defaults: /mnt/user-data/uploads  and  ./golden_out)

REQUIRES in uploads_dir (same names as window-3 uploads):
    Monthly_report.csv
    General_stock_--_Amazon_format__Jul_7__2026_.csv
    Stock_FC_wise_--_Amazon_format__Jul_7__2026_.csv
    fc_registration.pdf
    inventory_plan_template.xlsx            (the REAL master)
    daily_stock_workbook__Jul_7__2026_.xlsx (July-7 golden workbook, OLD labels)
plus configurations.xlsx (committed v2.8 version) next to this script or in
uploads_dir, and ingestion_v2 / workbook_builder_v2 / calculation_v2 on path.

THE THREE FIXTURES (regenerated fresh on every run, never committed):
  F1 registration CSV = the 40 real PDF pairs + MAA4,TN  (spec §5 exception
     path, exactly how a user resolves the GST-suppressed Chennai FC).
  F2 config copy with freshness_tolerance_days -> 10 ONLY (sample sales end
     30 Jun vs stock 05 Jul = 5-day gap; production stays 3).
  F3 label-translated copy of the July-7 workbook: OLD plain labels ->
     current registration-derived labels, + a Pune (MH) lead row of 10
     (COPIED FROM BOMBAY — a fixture guess flagged in the window-3 handover;
     confirm with the user, discovery D1).

GOLDEN CHECKS: fill the GOLDEN dict, rerun. Empty dict = dump-only mode.
Remember discovery D2: ses-500 Tier-2 currently proposes ceil(raw) (17 at
Bangalore), NOT the old app's inner-pack 15 — assumption A1 must be resolved
with the user BEFORE "fixing" either number.
"""

from __future__ import annotations

import csv
import re
import shutil
import sys
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")

import pandas as pd
from openpyxl import load_workbook

import ingestion_v2 as ing
from workbook_builder_v2 import read_stock_workbook
from calculation_v2 import run_calculation, detect_asin_ambiguities

# --------------------------------------------------------------- parameters

DAYS_OF_COVER = 30            # window-3 smoke value; change per test scenario
ASIN_JUDGMENTS: dict = {}     # e.g. {"B0XXXXXXXX": "only_active"}

# Fill with known-good values, then rerun. Keys are examples of the intended
# granularity; every filled entry is asserted, everything else just dumps.
GOLDEN = {
    # ("SKU_U", "Region label"): {"rounded_qty": 0, "raw_requirement": 0.0},
    # ("SES-500", "Bangalore (KA)"): {"rounded_qty": 15},   # <- ONLY after A1
    #                                                       #    is resolved!
    "units_planned_total": None,        # int, or None to skip
    "demand_units_resolved": 2925.0,    # window-3 verified: full coverage
    "demand_units_excluded": 0.0,       # window-3 verified: zero excluded
}

OLD_TO_NEW_LABELS = {   # window-3 discovery D1: July-7 workbook label style
    "DELHI": "Delhi (DL)", "HARYANA": "Haryana (HR)", "BOMBAY": "Bombay (MH)",
    "BANGALORE": "Bangalore (KA)", "HYDERABAD": "Hyderabad (TG)",
    "CHENNAI": "Chennai (TN)", "CALCUTTA": "Calcutta (WB)",
    "LUCKNOW": "Lucknow (UP)",
}
PUNE_LEAD_DAYS = 10     # fixture guess = Bombay's lead; CONFIRM WITH USER (D1)


# ----------------------------------------------------------------- fixtures

def make_registration_csv(uploads: Path, out: Path) -> Path:
    """F1: real PDF pairs + MAA4->TN (§5 exception, user-confirmed)."""
    import pdfplumber
    with pdfplumber.open(uploads / "fc_registration.pdf") as pdf:
        text = "\n".join(p.extract_text() or "" for p in pdf.pages)
    reg = {}
    for st, fc in re.findall(r"\b([A-Z]{2})\s+([A-Z0-9]{3,5})\s+\d", text):
        reg.setdefault(fc, st)
    if len(reg) != 40:
        print(f"  [!] registration parse found {len(reg)} FCs (expected 40) "
              f"— PDF layout may have changed; inspect before trusting.")
    reg["MAA4"] = "TN"
    path = out / "FIXTURE_fc_registration.csv"
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["State", "Fulfillment Center"])
        for fc, st in sorted(reg.items()):
            w.writerow([st, fc])
    return path


def make_config_copy(config_src: Path, out: Path) -> Path:
    """F2: freshness tolerance 3 -> 10, nothing else touched."""
    path = out / "FIXTURE_configurations.xlsx"
    shutil.copy(config_src, path)
    wb = load_workbook(path)
    ws = wb["Config"]
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == "freshness_tolerance_days":
            ws.cell(r, 2, 10)
    wb.save(path)
    return path


def make_translated_workbook(uploads: Path, out: Path) -> Path:
    """F3: OLD labels -> current labels in Reference / Current stock /
    In-transit / At the FC, plus the Pune (MH) lead row."""
    path = out / "FIXTURE_daily_stock_workbook.xlsx"
    shutil.copy(uploads / "daily_stock_workbook__Jul_7__2026_.xlsx", path)
    wb = load_workbook(path)

    ws, last = wb["Reference"], None
    for r in range(3, 40):
        v = ws.cell(r, 1).value
        if v in OLD_TO_NEW_LABELS:
            ws.cell(r, 1, OLD_TO_NEW_LABELS[v])
        if v:
            last = r
    already = any(ws.cell(r, 1).value == "Pune (MH)" for r in range(3, 40))
    if not already:
        ws.cell(last + 1, 1, "Pune (MH)")
        ws.cell(last + 1, 2, PUNE_LEAD_DAYS)

    cs = wb["Current stock"]
    for c in range(1, cs.max_column + 1):
        v = cs.cell(2, c).value
        if v in OLD_TO_NEW_LABELS:
            cs.cell(2, c, OLD_TO_NEW_LABELS[v])

    for tab in ("In-transit", "At the FC"):
        t = wb[tab]
        for r in range(1, t.max_row + 1):
            for c in range(1, 4):
                v = t.cell(r, c).value
                if isinstance(v, str) and v.strip().upper() in OLD_TO_NEW_LABELS:
                    t.cell(r, c, OLD_TO_NEW_LABELS[v.strip().upper()])
    wb.save(path)
    return path


# ------------------------------------------------------------------ pipeline

def run(uploads: Path, out: Path) -> int:
    out.mkdir(parents=True, exist_ok=True)
    fails: list[str] = []

    config_src = Path("configurations.xlsx")
    if not config_src.exists():
        config_src = uploads / "configurations.xlsx"
    if not config_src.exists():
        print("FATAL: configurations.xlsx not found next to script or in "
              "uploads dir."); return 2

    print("== fixtures ==")
    # v2.9: F1 (registration CSV) and F2 (tolerance copy) are OBSOLETE —
    # the FC-review gate takes MAA4 via fc_resolutions against the REAL PDF,
    # and the 5-day sales/stock gap passes silently (< sales_recency_prompt_days).
    f3 = make_translated_workbook(uploads, out)
    print(f"  F3 {f3.name} (Pune lead = {PUNE_LEAD_DAYS}, CONFIRM) — only fixture left")

    print("== ingestion ==")
    can, rep = ing.run_ingestion(
        open(uploads / "Monthly_report.csv", "rb"),
        open(uploads / "General_stock_--_Amazon_format__Jul_7__2026_.csv", "rb"),
        open(uploads / "Stock_FC_wise_--_Amazon_format__Jul_7__2026_.csv", "rb"),
        config_path=str(config_src),
        master_path=str(uploads / "inventory_plan_template.xlsx"),
        fcreg_path=str(uploads / "fc_registration.pdf"),
        fc_resolutions={"MAA4": {"action": "map",
                                 "region": "Chennai (TN)",
                                 "fulfillable": True}})   # v2.9 §5c
    for e in rep.errors:
        print("  ERR:", e)
    if not rep.ok:
        print("FATAL: ingestion failed."); return 2

    print("== workbook read ==")
    stock, rep2 = read_stock_workbook(str(f3), can)
    for e in rep2.errors[:10]:
        print("  ERR:", e)
    if not rep2.ok:
        print("FATAL: workbook read failed."); return 2

    amb = detect_asin_ambiguities(can, stock)
    if amb and not all(a in ASIN_JUDGMENTS for a in amb):
        print(f"  [!] unjudged ASIN ambiguities {amb} — engine defaults to "
              f"COMBINE with a warning; set ASIN_JUDGMENTS to control.")

    print("== calculation ==")
    res = run_calculation(can, stock, DAYS_OF_COVER,
                          asin_judgments=ASIN_JUDGMENTS or None)
    m = res.meta
    print(f"  plan {res.plan.shape} | planned {m['units_planned_total']}u | "
          f"resolved {m['demand_units_resolved']:.0f}u | "
          f"excluded {m['demand_units_excluded']:.0f}u | "
          f"stockouts {len(res.stockouts)} | warnings {len(res.warnings)}")
    for w in res.warnings:
        print("  WARN:", w)

    # ---- dump everything a human comparison needs
    dump = out / "golden_comparison.xlsx"
    with pd.ExcelWriter(dump) as xl:
        res.plan.to_excel(xl, sheet_name="plan", index=False)
        res.stockouts.to_excel(xl, sheet_name="stockouts", index=False)
        (res.plan.pivot_table(index="sku_u", columns="region",
                              values="rounded_qty", aggfunc="sum",
                              fill_value=0)
         .reindex(columns=[r for r in m["region_order"]
                           if r in res.plan["region"].unique()])
         .to_excel(xl, sheet_name="rounded_pivot"))
        pd.DataFrame([{"region": k, "priority": v}
                      for k, v in res.region_priorities.items()]
                     ).to_excel(xl, sheet_name="region_priorities", index=False)
        pd.DataFrame([{"key": k, "value": str(v)} for k, v in m.items()]
                     ).to_excel(xl, sheet_name="meta", index=False)
    print(f"  dumped -> {dump}")

    # ---- golden assertions (only filled entries)
    print("== golden checks ==")
    idx = res.plan.set_index(["sku_u", "region"])
    for key, expect in GOLDEN.items():
        if expect is None:
            continue
        if isinstance(key, tuple):
            if key not in idx.index:
                fails.append(f"{key}: row absent from plan"); continue
            row = idx.loc[key]
            for col, want in expect.items():
                got = row[col]
                if isinstance(want, float):
                    ok = abs(float(got) - want) < 0.05
                else:
                    ok = got == want
                (print if ok else fails.append)(
                    f"  {'OK ' if ok else ''}{key} {col}: got {got} "
                    f"want {want}" if ok else
                    f"{key} {col}: got {got}, want {want}")
        else:
            got = m.get(key)
            ok = (abs(float(got) - float(expect)) < 0.05
                  if isinstance(expect, float) else got == expect)
            (print if ok else fails.append)(
                f"  OK  meta.{key} = {got}" if ok else
                f"meta.{key}: got {got}, want {expect}")

    filled = [k for k, v in GOLDEN.items() if v is not None]
    if not filled:
        print("  (GOLDEN empty — dump-only mode; fill it and rerun)")
    if fails:
        print(f"\n== RESULT: {len(fails)} GOLDEN FAILURE(S) ==")
        for f_ in fails:
            print("  FAIL:", f_)
        return 1
    print("\n== RESULT: PASS "
          f"({len(filled)} golden entries checked) ==")
    return 0


if __name__ == "__main__":
    uploads = Path(sys.argv[1]) if len(sys.argv) > 1 \
        else Path("/mnt/user-data/uploads")
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("golden_out")
    sys.exit(run(uploads, out))
