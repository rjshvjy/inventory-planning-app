import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import io
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')

# Page configuration
st.set_page_config(
    page_title="Inventory Planning Tool",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Initialize session state
if 'files_uploaded' not in st.session_state:
    st.session_state.files_uploaded = False
if 'inventory_data' not in st.session_state:
    st.session_state.inventory_data = {}
if 'file_mapping' not in st.session_state:
    st.session_state.file_mapping = {}
if 'master_sku_list' not in st.session_state:
    st.session_state.master_sku_list = []
if 'calculation_complete' not in st.session_state:
    st.session_state.calculation_complete = False

# Title and header
st.title("🏭 Inventory Planning Tool")
st.markdown("---")

# Sidebar for file uploads
with st.sidebar:
    st.header("📁 File Upload")
    
    # File uploaders
    monthly_sales_file = st.file_uploader(
        "Monthly Sales Report", 
        type=['xlsx', 'xls'],
        help="Upload the monthly sales report with SKUs and state-wise sales"
    )
    
    states_mapping_file = st.file_uploader(
        "States Mapping File", 
        type=['xlsx', 'xls'],
        help="Upload the states mapping file with FC locations and delivery times"
    )
    
    daily_stock_file = st.file_uploader(
        "Daily Stock Report", 
        type=['xlsx', 'xls'],
        help="Upload the daily stock report with Overall stock and In-transit tabs"
    )
    
    inventory_plan_template = st.file_uploader(
        "Inventory Plan Template (Optional)", 
        type=['xlsx', 'xls'],
        help="Upload the inventory plan template to fill with calculated values"
    )
    
    # Target stock days input
    st.markdown("---")
    target_stock_days = st.number_input(
        "Target Stock Days",
        min_value=1,
        max_value=90,
        value=15,
        help="Number of days of sales worth of stock to maintain at each location"
    )
    
    # Process button
    if st.button("🚀 Process Files", type="primary"):
        if monthly_sales_file and states_mapping_file and daily_stock_file:
            st.session_state.files_uploaded = True
        else:
            st.error("Please upload all required files!")

# Helper functions
def identify_file_type(filename):
    """Identify file type based on filename"""
    filename_lower = filename.lower()
    name_without_ext = filename_lower.replace('.xlsx', '').replace('.xls', '').replace('.csv', '')
    
    if any(keyword in filename_lower for keyword in ['monthly', 'month']) and \
       any(keyword in filename_lower for keyword in ['sales', 'sale']):
        return 'monthly_sales'
    
    if 'state' in filename_lower and any(keyword in filename_lower for keyword in ['map', 'mapping']):
        return 'states_mapping'
    elif 'states' in filename_lower:
        return 'states_mapping'
    
    if 'daily' in filename_lower:
        return 'daily_stock'
    elif 'stock' in filename_lower and any(keyword in filename_lower for keyword in ['report', 'daily', 'transit']):
        return 'daily_stock'
    
    if 'inventory' in filename_lower and 'plan' in filename_lower:
        return 'inventory_plan'
    
    return 'unknown'

def extract_skus_from_monthly_sales(df_monthly):
    """Extract SKUs from monthly sales report"""
    # Find the row containing 'SKU'
    sku_row = None
    for idx, row in df_monthly.iterrows():
        if any('SKU' in str(cell) for cell in row):
            sku_row = idx
            break
    
    if sku_row is not None:
        # Read again with proper header
        df_monthly = pd.read_excel(monthly_sales_file, header=sku_row)
        
        # The SKU column should be the first column
        sku_col = df_monthly.columns[0]
        
        # Extract SKUs, excluding 'Total' rows
        state_col = df_monthly.columns[1]  # Usually 'Row Labels'
        df_filtered = df_monthly[~df_monthly[state_col].astype(str).str.contains('Total|total', case=False, na=False)]
        
        skus = df_filtered[sku_col].dropna().astype(str).str.strip()
        skus = skus[skus != ''].str.upper().unique()
        
        return list(skus), df_monthly
    
    return [], df_monthly

def load_and_process_data(monthly_sales_file, states_mapping_file, daily_stock_file, target_stock_days):
    """Load and process all data files"""
    inventory_data = {}
    
    with st.spinner("Loading and processing files..."):
        progress_bar = st.progress(0)
        
        # 1. Load Monthly Sales
        progress_bar.progress(10)
        st.info("Loading monthly sales data...")
        
        df_monthly_raw = pd.read_excel(monthly_sales_file, header=None)
        
        # Find the row with actual data
        data_start_row = None
        for i in range(len(df_monthly_raw)):
            if any(keyword in str(df_monthly_raw.iloc[i].values) for keyword in ['Puvi', 'SKU']):
                if 'SKU' in str(df_monthly_raw.iloc[i].values):
                    data_start_row = i
                    break
        
        # Read with correct header
        df_monthly = pd.read_excel(monthly_sales_file, header=data_start_row)
        
        # Clean column names
        df_monthly.columns = df_monthly.columns.str.strip()
        
        # Rename columns if needed
        if 'Unnamed: 0' in df_monthly.columns:
            df_monthly.rename(columns={'Unnamed: 0': 'SKU'}, inplace=True)
        
        col_names = list(df_monthly.columns)
        if len(col_names) > 1 and 'Row Labels' not in col_names:
            df_monthly.rename(columns={col_names[1]: 'Row Labels'}, inplace=True)
        
        # Get state columns
        state_columns = []
        for col in df_monthly.columns[2:]:
            if col not in ['Grand Total', 'Unnamed: 36'] and not str(col).startswith('Unnamed'):
                state_columns.append(col)
        
        inventory_data['monthly_sales'] = df_monthly
        inventory_data['state_columns'] = state_columns
        
        # Extract SKUs
        master_sku_list = []
        for idx, row in df_monthly.iterrows():
            sku = str(row['SKU']).upper()
            if (sku not in ['NAN', 'NONE', ''] and 
                pd.notna(row['SKU']) and 
                len(sku) >= 3 and
                'total' not in str(row.get('Row Labels', '')).lower()):
                master_sku_list.append(sku)
        
        st.session_state.master_sku_list = list(set(master_sku_list))
        
        # 2. Load States Mapping
        progress_bar.progress(30)
        st.info("Loading states mapping data...")
        
        # Load all sheets
        df_states_map = pd.read_excel(states_mapping_file, sheet_name='StatesMapping')
        df_fc_location = pd.read_excel(states_mapping_file, sheet_name='FC-Location')
        df_delivery = pd.read_excel(states_mapping_file, sheet_name='DeliveryTime')
        
        inventory_data['state_mapping'] = df_states_map
        inventory_data['fc_location'] = df_fc_location
        inventory_data['delivery_times'] = df_delivery
        
        # 3. Load Current Stock
        progress_bar.progress(50)
        st.info("Loading current stock data...")
        
        df_stock = pd.read_excel(daily_stock_file, sheet_name='Overall stock', header=1)
        
        # Clean SKU column
        df_stock = df_stock[df_stock['SKU'].notna()]
        df_stock = df_stock[~df_stock['Row Labels'].str.contains('In transit -', case=False, na=False)]
        
        # Get location columns
        location_columns = [col for col in df_stock.columns[2:]
                           if 'Unnamed' not in str(col)
                           and col not in ['Grand Total', 'CURRENT', 'STOCK']]
        
        inventory_data['current_stock'] = df_stock
        inventory_data['location_columns'] = location_columns
        
        # Check for At the FC stock
        try:
            df_at_fc = pd.read_excel(daily_stock_file, sheet_name='At the FC', header=1)
            df_at_fc = df_at_fc[df_at_fc.iloc[:, 0].notna()]
            inventory_data['at_fc'] = df_at_fc
        except:
            inventory_data['at_fc'] = pd.DataFrame()
        
        # 4. Load In-Transit
        progress_bar.progress(70)
        st.info("Loading in-transit data...")
        
        try:
            df_transit = pd.read_excel(daily_stock_file, sheet_name='In-transit')
            df_transit = df_transit.dropna(how='all')
            df_transit = df_transit[df_transit['SKU'].notna()]
            
            if len(df_transit) > 0:
                df_transit['Receiving date'] = pd.to_datetime(df_transit['Receiving date'], errors='coerce')
                df_transit['Days until arrival'] = (df_transit['Receiving date'] - datetime.now()).dt.days
            
            inventory_data['in_transit'] = df_transit
        except:
            inventory_data['in_transit'] = pd.DataFrame()
        
        progress_bar.progress(100)
        st.success("All files loaded successfully!")
        
    return inventory_data

def create_state_to_location_mapping(inventory_data):
    """Create mapping from states to warehouse locations"""
    state_to_location_map = {}
    
    # Create FC to Location mapping
    fc_to_location = {}
    for idx, row in inventory_data['fc_location'].iterrows():
        fc_code = str(row.iloc[0]).strip().upper()
        location = str(row.iloc[1]).strip().upper()
        fc_to_location[fc_code] = location
    
    # Load states mapping
    df_states_mapping = inventory_data['state_mapping']
    
    # Try to extract state mappings
    state_to_fc = {}
    
    # Check structure and extract mappings
    for idx, row in df_states_mapping.iterrows():
        for col_idx in range(len(row)):
            for fc_col_idx in range(len(row)):
                if col_idx != fc_col_idx:
                    state = str(row.iloc[col_idx]).strip() if pd.notna(row.iloc[col_idx]) else None
                    fc = str(row.iloc[fc_col_idx]).strip().upper() if pd.notna(row.iloc[fc_col_idx]) else None
                    
                    if state and fc and fc in fc_to_location:
                        if any(c.isalpha() for c in state) and state in inventory_data.get('state_columns', []):
                            state_to_fc[state] = fc
                            break
    
    # Create final mapping
    for state in inventory_data.get('state_columns', []):
        state_clean = state.strip()
        mapped = False
        
        # Try exact match
        if state_clean in state_to_fc:
            fc = state_to_fc[state_clean]
            if fc in fc_to_location:
                location = fc_to_location[fc]
                state_to_location_map[state] = {
                    'fc': fc,
                    'location': location
                }
                mapped = True
        
        # Default mapping if not found
        if not mapped:
            # Geographic default mappings
            default_mappings = {
                'Andhra Pradesh': 'HYDERABAD',
                'Delhi': 'DELHI',
                'Maharashtra': 'BOMBAY',
                'Karnataka': 'BANGALORE',
                'Tamil Nadu': 'CHENNAI',
                'West Bengal': 'CALCUTTA',
                'Uttar Pradesh': 'LUCKNOW',
                'Haryana': 'HARYANA'
            }
            
            default_location = default_mappings.get(state, 'DELHI')
            default_fc = 'DEX8'
            
            for fc, loc in fc_to_location.items():
                if loc == default_location:
                    default_fc = fc
                    break
            
            state_to_location_map[state] = {
                'fc': default_fc,
                'location': default_location
            }
    
    return state_to_location_map, fc_to_location

def calculate_location_wise_sales(inventory_data, state_to_location_map):
    """Calculate daily sales by location"""
    days_in_month = 30  # Assuming 30 days
    
    location_daily_sales = {}
    sku_location_sales = {}
    
    df_monthly = inventory_data['monthly_sales']
    
    for idx, row in df_monthly.iterrows():
        sku = str(row['SKU']).upper()
        
        # Skip invalid SKUs
        if sku in ['NAN', 'NONE', ''] or pd.isna(row['SKU']) or len(sku) < 3:
            continue
        
        # Skip total rows
        if 'total' in str(row.get('Row Labels', '')).lower():
            continue
        
        product_name = row['Row Labels']
        sku_sales_by_location = {}
        
        # Calculate sales by location
        for state in inventory_data['state_columns']:
            monthly_qty = row[state] if pd.notna(row[state]) else 0
            
            if monthly_qty > 0 and state in state_to_location_map:
                location = state_to_location_map[state]['location']
                
                if location not in sku_sales_by_location:
                    sku_sales_by_location[location] = 0
                sku_sales_by_location[location] += monthly_qty
        
        # Convert to daily sales
        for location, monthly_qty in sku_sales_by_location.items():
            daily_sales = monthly_qty / days_in_month
            
            if sku not in sku_location_sales:
                sku_location_sales[sku] = {}
            
            sku_location_sales[sku][location] = {
                'monthly_sales': monthly_qty,
                'daily_sales': daily_sales,
                'product_name': product_name
            }
            
            if location not in location_daily_sales:
                location_daily_sales[location] = 0
            location_daily_sales[location] += daily_sales
    
    return location_daily_sales, sku_location_sales

def analyze_current_stock(inventory_data):
    """Analyze current stock levels (Cell 14 V3 logic)"""
    current_stock_by_sku_location = {}
    in_transit_from_overall = {}
    at_fc_from_overall = {}
    
    current_sku = None
    
    # Process Overall stock sheet with hierarchical structure
    for idx, row in inventory_data['current_stock'].iterrows():
        sku_cell = row['SKU']
        row_label = str(row['Row Labels']) if pd.notna(row['Row Labels']) else ""
        
        # Check if this is a main SKU row
        if pd.notna(sku_cell) and str(sku_cell).strip() and str(sku_cell).upper() != 'NAN':
            current_sku = str(sku_cell).strip().upper()
            
            if current_sku not in current_stock_by_sku_location:
                current_stock_by_sku_location[current_sku] = {}
            
            # Process each location
            for location in inventory_data['location_columns']:
                if location.upper() == 'YSXA':
                    continue
                
                stock_value = row[location]
                
                # Convert to float safely
                if pd.isna(stock_value) or stock_value == '' or stock_value is None:
                    stock_qty = 0
                else:
                    try:
                        if isinstance(stock_value, str) and any(keyword in str(stock_value) for keyword in ['Puvi', 'Oil', 'Litre', 'Pressed']):
                            stock_qty = 0
                        else:
                            stock_qty = float(stock_value)
                            stock_qty = max(0, stock_qty)
                    except (ValueError, TypeError):
                        stock_qty = 0
                
                current_stock_by_sku_location[current_sku][location] = stock_qty
        
        # Check for sub-rows
        elif current_sku and 'In transit -' in row_label:
            if current_sku not in in_transit_from_overall:
                in_transit_from_overall[current_sku] = {}
            
            for location in inventory_data['location_columns']:
                if location.upper() == 'YSXA':
                    continue
                
                transit_value = row[location]
                if pd.notna(transit_value) and transit_value != '':
                    try:
                        transit_qty = float(transit_value)
                        if transit_qty > 0:
                            if location not in in_transit_from_overall[current_sku]:
                                in_transit_from_overall[current_sku][location] = 0
                            in_transit_from_overall[current_sku][location] += transit_qty
                    except (ValueError, TypeError):
                        pass
        
        elif current_sku and 'At the FC -' in row_label:
            if current_sku not in at_fc_from_overall:
                at_fc_from_overall[current_sku] = {}
            
            for location in inventory_data['location_columns']:
                if location.upper() == 'YSXA':
                    continue
                
                fc_value = row[location]
                if pd.notna(fc_value) and fc_value != '':
                    try:
                        fc_qty = float(fc_value)
                        if fc_qty > 0:
                            if location not in at_fc_from_overall[current_sku]:
                                at_fc_from_overall[current_sku][location] = 0
                            at_fc_from_overall[current_sku][location] += fc_qty
                    except (ValueError, TypeError):
                        pass
    
    # Add at-FC quantities to current stock
    for sku, locations in at_fc_from_overall.items():
        for location, qty in locations.items():
            if sku not in current_stock_by_sku_location:
                current_stock_by_sku_location[sku] = {}
            if location not in current_stock_by_sku_location[sku]:
                current_stock_by_sku_location[sku][location] = 0
            
            current_stock_by_sku_location[sku][location] += qty
    
    # Process separate "At the FC" sheet if exists
    if 'at_fc' in inventory_data and len(inventory_data['at_fc']) > 0:
        for idx, row in inventory_data['at_fc'].iterrows():
            # Extract SKU
            sku = None
            if 'SKU' in row.index and pd.notna(row['SKU']):
                sku = str(row['SKU']).strip().upper()
            elif pd.notna(row.iloc[0]):
                potential_sku = str(row.iloc[0]).strip()
                if len(potential_sku) < 30:
                    sku = potential_sku.upper()
            
            if sku and sku != 'NAN':
                if sku not in current_stock_by_sku_location:
                    current_stock_by_sku_location[sku] = {}
                
                for location in inventory_data['location_columns']:
                    if location.upper() == 'YSXA':
                        continue
                    
                    if location in row.index:
                        qty_value = row[location]
                        if pd.notna(qty_value) and qty_value != '':
                            try:
                                qty = float(qty_value)
                                if qty > 0:
                                    if location not in current_stock_by_sku_location[sku]:
                                        current_stock_by_sku_location[sku][location] = 0
                                    current_stock_by_sku_location[sku][location] += qty
                            except (ValueError, TypeError):
                                pass
    
    return current_stock_by_sku_location, in_transit_from_overall, at_fc_from_overall

def process_lead_times_and_transit(inventory_data, state_to_location_map, in_transit_from_overall):
    """Process lead times and in-transit inventory (Cell 15 logic)"""
    # Load delivery lead times
    delivery_lead_times = {}
    df_delivery = inventory_data['delivery_times']
    
    for idx, row in df_delivery.iterrows():
        fc_code = str(row.iloc[0]).strip()
        lead_days = int(row.iloc[1]) if pd.notna(row.iloc[1]) else 3
        delivery_lead_times[fc_code] = lead_days
    
    # Map to locations
    location_lead_times = {}
    for state, mapping in state_to_location_map.items():
        fc = mapping['fc']
        location = mapping['location']
        if fc in delivery_lead_times:
            location_lead_times[location] = delivery_lead_times[fc]
    
    # Process in-transit inventory from both sources
    in_transit_by_sku_location = {}
    in_transit_details = []
    
    # Add in-transit from Overall stock sheet
    for sku, locations in in_transit_from_overall.items():
        for location, quantity in locations.items():
            if quantity > 0:
                lead_time = location_lead_times.get(location, 3)
                receiving_date = datetime.now() + timedelta(days=lead_time)
                
                if sku not in in_transit_by_sku_location:
                    in_transit_by_sku_location[sku] = {}
                if location not in in_transit_by_sku_location[sku]:
                    in_transit_by_sku_location[sku][location] = []
                
                transit_info = {
                    'quantity': quantity,
                    'receiving_date': receiving_date,
                    'days_until_arrival': lead_time,
                    'source': 'Overall stock sheet'
                }
                
                in_transit_by_sku_location[sku][location].append(transit_info)
                in_transit_details.append({
                    'sku': sku,
                    'location': location,
                    'quantity': quantity,
                    'receiving_date': receiving_date,
                    'days_until_arrival': lead_time,
                    'source': 'Overall stock sheet'
                })
    
    # Add from In-transit sheet
    if 'in_transit' in inventory_data and len(inventory_data['in_transit']) > 0:
        for idx, row in inventory_data['in_transit'].iterrows():
            sku = str(row['SKU']).upper()
            location = row['Location']
            quantity = row['Quantity'] if pd.notna(row['Quantity']) else 0
            receiving_date = row['Receiving date']
            
            if pd.notna(receiving_date) and quantity > 0:
                days_until_arrival = (receiving_date - datetime.now()).days
                
                if sku not in in_transit_by_sku_location:
                    in_transit_by_sku_location[sku] = {}
                if location not in in_transit_by_sku_location[sku]:
                    in_transit_by_sku_location[sku][location] = []
                
                transit_info = {
                    'quantity': quantity,
                    'receiving_date': receiving_date,
                    'days_until_arrival': days_until_arrival,
                    'source': 'In-transit sheet'
                }
                
                in_transit_by_sku_location[sku][location].append(transit_info)
                in_transit_details.append({
                    'sku': sku,
                    'location': location,
                    'quantity': quantity,
                    'receiving_date': receiving_date,
                    'days_until_arrival': days_until_arrival,
                    'source': 'In-transit sheet'
                })
    
    return location_lead_times, in_transit_by_sku_location, in_transit_details

def calculate_shipment_requirements(master_sku_list, inventory_data, sku_location_sales, 
                                  current_stock_by_sku_location, location_lead_times, 
                                  in_transit_by_sku_location, target_stock_days):
    """Calculate shipment requirements (Cell 16 logic)"""
    shipment_requirements = {}
    stockout_warnings = []
    
    for sku in master_sku_list:
        sku_upper = sku.upper()
        shipment_requirements[sku_upper] = {}
        
        # Get product name
        product_name = None
        for idx, row in inventory_data['current_stock'].iterrows():
            if str(row['SKU']).upper() == sku_upper:
                product_name = row['Row Labels']
                break
        
        if not product_name:
            for idx, row in inventory_data['monthly_sales'].iterrows():
                if str(row['SKU']).upper() == sku_upper:
                    product_name = row['Row Labels']
                    break
        
        # For each location
        for location in inventory_data['location_columns']:
            if location.upper() == 'YSXA':
                continue
            
            # Get daily sales
            daily_sales = 0
            if sku_upper in sku_location_sales and location in sku_location_sales[sku_upper]:
                daily_sales = sku_location_sales[sku_upper][location]['daily_sales']
            
            # Get lead time
            lead_time = location_lead_times.get(location, 3)
            
            # Get current stock
            current_stock = 0
            if sku_upper in current_stock_by_sku_location and location in current_stock_by_sku_location[sku_upper]:
                current_stock = current_stock_by_sku_location[sku_upper][location]
            
            # Calculate requirements
            today = 0
            shipment_arrival_day = lead_time
            target_end_day = shipment_arrival_day + target_stock_days
            
            sales_during_lead_time = daily_sales * lead_time
            sales_during_target_period = daily_sales * target_stock_days
            total_depletion = sales_during_lead_time + sales_during_target_period
            
            available_stock = current_stock
            
            # Add in-transit stock
            if sku_upper in in_transit_by_sku_location and location in in_transit_by_sku_location[sku_upper]:
                for transit in in_transit_by_sku_location[sku_upper][location]:
                    if transit['days_until_arrival'] <= target_end_day:
                        available_stock += transit['quantity']
            
            stock_gap = total_depletion - available_stock
            required_shipment = max(0, round(stock_gap))
            
            # Stockout detection
            if daily_sales > 0:
                running_stock = current_stock
                for day in range(target_end_day + 1):
                    running_stock -= daily_sales
                    
                    if sku_upper in in_transit_by_sku_location and location in in_transit_by_sku_location[sku_upper]:
                        for transit in in_transit_by_sku_location[sku_upper][location]:
                            if transit['days_until_arrival'] == day:
                                running_stock += transit['quantity']
                    
                    if day == shipment_arrival_day and required_shipment > 0:
                        running_stock += required_shipment
                    
                    if running_stock < 0 and day < target_end_day:
                        stockout_warnings.append({
                            'sku': sku_upper,
                            'product': product_name,
                            'location': location,
                            'stockout_day': day,
                            'deficit': abs(running_stock)
                        })
                        break
            
            # Store requirement
            shipment_requirements[sku_upper][location] = {
                'required_shipment': required_shipment,
                'daily_sales': daily_sales,
                'current_stock': current_stock,
                'lead_time': lead_time,
                'product_name': product_name,
                'calculation_details': {
                    'sales_during_lead_time': sales_during_lead_time,
                    'sales_during_target_period': sales_during_target_period,
                    'total_depletion': total_depletion,
                    'available_stock': available_stock,
                    'stock_gap': stock_gap
                }
            }
    
    return shipment_requirements, stockout_warnings

def calculate_location_priorities(shipment_requirements):
    """Calculate location priorities based on risk scores"""
    location_risk_scores = {}
    sku_location_priorities = {}
    
    for sku, locations in shipment_requirements.items():
        sku_location_priorities[sku] = {}
        
        for location, details in locations.items():
            if details['daily_sales'] > 0:
                current_stock = details['current_stock']
                daily_sales = details['daily_sales']
                days_of_stock = current_stock / daily_sales if daily_sales > 0 else 999
                
                lead_time = details['lead_time']
                risk_score = lead_time - days_of_stock
                
                if risk_score >= 2 or current_stock == 0:
                    priority = 'Pri-1'
                elif risk_score >= 0:
                    priority = 'Pri-2'
                elif risk_score >= -3:
                    priority = 'Pri-3'
                else:
                    priority = 'Pri-4'
                
                sku_location_priorities[sku][location] = {
                    'priority': priority,
                    'risk_score': risk_score,
                    'days_of_stock': days_of_stock,
                    'required_shipment': details['required_shipment']
                }
                
                if location not in location_risk_scores:
                    location_risk_scores[location] = []
                location_risk_scores[location].append(risk_score)
    
    # Calculate average risk per location
    location_priorities = {}
    for location, risk_scores in location_risk_scores.items():
        avg_risk = sum(risk_scores) / len(risk_scores) if risk_scores else 0
        
        if avg_risk >= 1:
            location_priorities[location] = 'Pri-1'
        elif avg_risk >= -1:
            location_priorities[location] = 'Pri-2'
        elif avg_risk >= -3:
            location_priorities[location] = 'Pri-3'
        else:
            location_priorities[location] = 'Pri-4'
    
    return location_priorities, sku_location_priorities

def generate_excel_output(shipment_requirements, location_priorities, inventory_data, 
                         stockout_warnings, target_stock_days, inventory_plan_template=None):
    """Generate Excel output with all sheets"""
    output = io.BytesIO()
    
    if inventory_plan_template:
        # Use template
        wb = load_workbook(inventory_plan_template, keep_vba=False, data_only=False)
        ws = wb.active
        
        # Read location mappings from template
        location_to_column = {}
        for col in range(4, 12):  # D=4 to K=11
            cell_value = ws.cell(row=2, column=col).value
            if cell_value:
                location_clean = str(cell_value).strip().upper()
                location_to_column[location_clean] = col
        
        # Create SKU to row mapping
        sku_to_row = {}
        for row in range(3, ws.max_row + 1):
            sku_cell = ws.cell(row=row, column=1).value
            if sku_cell and str(sku_cell).strip():
                sku_clean = str(sku_cell).strip().upper()
                sku_to_row[sku_clean] = row
        
        # Fill in shipment requirements
        for sku, locations in shipment_requirements.items():
            sku_upper = sku.upper()
            
            if sku_upper in sku_to_row:
                row_num = sku_to_row[sku_upper]
                
                # Clear existing values
                for col in range(4, 12):
                    ws.cell(row=row_num, column=col).value = None
                
                # Fill new values
                for location, details in locations.items():
                    location_upper = location.upper()
                    
                    if location_upper in location_to_column:
                        col_num = location_to_column[location_upper]
                        qty = details['required_shipment']
                        
                        if qty > 0:
                            ws.cell(row=row_num, column=col_num).value = int(qty)
            else:
                # Add missing SKUs at the end
                next_row = ws.max_row + 1
                ws.cell(row=next_row, column=1).value = sku_upper
                
                # Add product name
                for loc_data in shipment_requirements[sku_upper].values():
                    if loc_data.get('product_name'):
                        ws.cell(row=next_row, column=2).value = loc_data['product_name']
                        break
                
                # Add quantities
                for location, details in shipment_requirements[sku_upper].items():
                    location_upper = location.upper()
                    if location_upper in location_to_column:
                        col_num = location_to_column[location_upper]
                        qty = details['required_shipment']
                        if qty > 0:
                            ws.cell(row=next_row, column=col_num).value = int(qty)
    else:
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Shipment_Plan"
        
        # Create main output data
        output_data = []
        
        for sku in sorted(shipment_requirements.keys()):
            row = {'SKU': sku}
            
            # Get product name
            for loc_data in shipment_requirements[sku].values():
                if loc_data.get('product_name'):
                    row['Product Name'] = loc_data['product_name']
                    break
            
            # Add quantities for each location
            for location in inventory_data['location_columns']:
                if location.upper() != 'YSXA':
                    priority = location_priorities.get(location, 'Pri-4')
                    col_name = f"{location} ({priority})"
                    
                    if location in shipment_requirements[sku]:
                        row[col_name] = shipment_requirements[sku][location]['required_shipment']
                    else:
                        row[col_name] = 0
            
            output_data.append(row)
        
        # Write to worksheet
        df_output = pd.DataFrame(output_data)
        for r_idx, row in enumerate(dataframe_to_rows(df_output, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Add Calculation Details sheet
    calc_ws = wb.create_sheet("Calculation_Details")
    calc_data = []
    
    for sku, locations in shipment_requirements.items():
        for location, details in locations.items():
            if details['daily_sales'] > 0 or details['required_shipment'] > 0:
                calc_data.append({
                    'SKU': sku,
                    'Product': details.get('product_name', ''),
                    'Location': location,
                    'Daily_Sales': round(details['daily_sales'], 1),
                    'Current_Stock': details['current_stock'],
                    'Lead_Time_Days': details['lead_time'],
                    'Target_Stock_Days': target_stock_days,
                    'Total_Days': details['lead_time'] + target_stock_days,
                    'Required_Units': details['required_shipment']
                })
    
    if calc_data:
        calc_df = pd.DataFrame(calc_data)
        for r_idx, row in enumerate(dataframe_to_rows(calc_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                calc_ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Add Summary sheet
    summary_ws = wb.create_sheet("Summary")
    
    total_units = sum(sum(loc['required_shipment'] for loc in locs.values()) 
                     for locs in shipment_requirements.values())
    
    summary_data = [
        ['Metric', 'Value'],
        ['Planning Date', datetime.now().strftime('%Y-%m-%d')],
        ['Target Stock Days', target_stock_days],
        ['Total SKUs', len(shipment_requirements)],
        ['Total Locations', len([loc for loc in inventory_data['location_columns'] if loc.upper() != 'YSXA'])],
        ['Total Units to Ship', total_units],
        ['Potential Stockouts', len(stockout_warnings)]
    ]
    
    for r_idx, row in enumerate(summary_data, 1):
        for c_idx, value in enumerate(row, 1):
            summary_ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Add Stockout Warnings sheet if any
    if stockout_warnings:
        warnings_ws = wb.create_sheet("Stockout_Warnings")
        
        warning_data = []
        for warning in stockout_warnings[:100]:
            warning_data.append({
                'SKU': warning['sku'],
                'Product': warning['product'][:50] if warning['product'] else '',
                'Location': warning['location'],
                'Stockout_Day': warning['stockout_day'],
                'Deficit_Units': warning['deficit']
            })
        
        if warning_data:
            warning_df = pd.DataFrame(warning_data)
            for r_idx, row in enumerate(dataframe_to_rows(warning_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    warnings_ws.cell(row=r_idx, column=c_idx, value=value)
    
    wb.save(output)
    output.seek(0)
    
    return output

# Main app logic
if st.session_state.files_uploaded:
    
    # Process data
    with st.spinner("Processing data..."):
        
        # Load all data
        inventory_data = load_and_process_data(
            monthly_sales_file, 
            states_mapping_file, 
            daily_stock_file,
            target_stock_days
        )
        st.session_state.inventory_data = inventory_data
        
        # Create state to location mapping
        state_to_location_map, fc_to_location = create_state_to_location_mapping(inventory_data)
        
        # Calculate location-wise sales
        location_daily_sales, sku_location_sales = calculate_location_wise_sales(
            inventory_data, state_to_location_map
        )
        
        # Analyze current stock (Cell 14 V3)
        current_stock_by_sku_location, in_transit_from_overall, at_fc_from_overall = analyze_current_stock(
            inventory_data
        )
        
        # Process lead times and in-transit (Cell 15)
        location_lead_times, in_transit_by_sku_location, in_transit_details = process_lead_times_and_transit(
            inventory_data, state_to_location_map, in_transit_from_overall
        )
        
        # Calculate shipment requirements (Cell 16)
        shipment_requirements, stockout_warnings = calculate_shipment_requirements(
            st.session_state.master_sku_list,
            inventory_data,
            sku_location_sales,
            current_stock_by_sku_location,
            location_lead_times,
            in_transit_by_sku_location,
            target_stock_days
        )
        
        # BLR4 IXD redistribution (if applicable)
        blr4_ixd_location = None
        for location in inventory_data['location_columns']:
            if 'BLR4' in location.upper() and 'IXD' in location.upper():
                blr4_ixd_location = location
                break
        
        if blr4_ixd_location:
            # Implement BLR4 IXD redistribution
            for sku in st.session_state.master_sku_list:
                sku_upper = sku.upper()
                
                # Get BLR4 IXD available stock
                blr4_stock = 0
                if sku_upper in current_stock_by_sku_location and blr4_ixd_location in current_stock_by_sku_location[sku_upper]:
                    blr4_stock = current_stock_by_sku_location[sku_upper][blr4_ixd_location]
                
                # Add in-transit to BLR4 IXD
                if sku_upper in in_transit_by_sku_location and blr4_ixd_location in in_transit_by_sku_location[sku_upper]:
                    for transit in in_transit_by_sku_location[sku_upper][blr4_ixd_location]:
                        if transit['days_until_arrival'] <= target_stock_days:
                            blr4_stock += transit['quantity']
                
                if blr4_stock > 0:
                    # Calculate total deficit across other locations
                    total_deficit = 0
                    deficit_locations = {}
                    
                    for location, details in shipment_requirements[sku_upper].items():
                        if location != blr4_ixd_location and details['required_shipment'] > 0:
                            deficit_locations[location] = {
                                'deficit': details['required_shipment'],
                                'daily_sales': details['daily_sales']
                            }
                            total_deficit += details['required_shipment']
                    
                    if total_deficit > 0 and len(deficit_locations) > 0:
                        # Distribute BLR4 IXD stock proportionally
                        total_daily_sales = sum(d['daily_sales'] for d in deficit_locations.values())
                        
                        if total_daily_sales > 0:
                            distribution_factor = min(1, blr4_stock / total_deficit)
                            
                            for location, deficit_info in deficit_locations.items():
                                sales_weight = deficit_info['daily_sales'] / total_daily_sales if total_daily_sales > 0 else 0
                                allocated_qty = min(
                                    deficit_info['deficit'],
                                    round(blr4_stock * sales_weight)
                                )
                                
                                if allocated_qty > 0:
                                    # Reduce the location's requirement
                                    shipment_requirements[sku_upper][location]['required_shipment'] -= allocated_qty
                                    shipment_requirements[sku_upper][location]['required_shipment'] = max(0,
                                        shipment_requirements[sku_upper][location]['required_shipment'])
        
        # Calculate priorities
        location_priorities, sku_location_priorities = calculate_location_priorities(
            shipment_requirements
        )
        
        st.session_state.calculation_complete = True
        st.session_state.shipment_requirements = shipment_requirements
        st.session_state.location_priorities = location_priorities
        st.session_state.stockout_warnings = stockout_warnings
        st.session_state.location_daily_sales = location_daily_sales
    
    # Display results
    st.success("✅ Processing complete!")
    
    # Summary metrics
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Total SKUs", len(st.session_state.master_sku_list))
    
    with col2:
        total_units = sum(sum(loc['required_shipment'] for loc in locs.values()) 
                         for locs in shipment_requirements.values())
        st.metric("Total Units to Ship", f"{total_units:,}")
    
    with col3:
        locations_needing_shipment = set()
        for sku, locs in shipment_requirements.items():
            for loc, details in locs.items():
                if details['required_shipment'] > 0:
                    locations_needing_shipment.add(loc)
        st.metric("Locations Needing Shipment", len(locations_needing_shipment))
    
    with col4:
        st.metric("Stockout Warnings", len(stockout_warnings))
    
    # Tabs for different views
    tab1, tab2, tab3, tab4 = st.tabs(["📊 Shipment Requirements", "⚠️ Stockout Warnings", 
                                       "📍 Location Summary", "📥 Download Results"])
    
    with tab1:
        st.subheader("Shipment Requirements by SKU and Location")
        
        # Create display dataframe
        display_data = []
        for sku in sorted(shipment_requirements.keys()):
            row = {'SKU': sku}
            
            # Get product name
            for loc_data in shipment_requirements[sku].values():
                if loc_data.get('product_name'):
                    row['Product'] = loc_data['product_name'][:50]
                    break
            
            # Add quantities
            for location in sorted(inventory_data['location_columns']):
                if location.upper() != 'YSXA':
                    if location in shipment_requirements[sku]:
                        row[location] = shipment_requirements[sku][location]['required_shipment']
                    else:
                        row[location] = 0
            
            display_data.append(row)
        
        df_display = pd.DataFrame(display_data)
        st.dataframe(df_display, use_container_width=True)
    
    with tab2:
        st.subheader("Stockout Warnings")
        
        if stockout_warnings:
            # Group by location
            warnings_by_location = {}
            for warning in stockout_warnings:
                loc = warning['location']
                if loc not in warnings_by_location:
                    warnings_by_location[loc] = []
                warnings_by_location[loc].append(warning)
            
            # Display by location
            for location, warnings in sorted(warnings_by_location.items()):
                with st.expander(f"📍 {location} ({len(warnings)} warnings)"):
                    for w in warnings[:10]:
                        st.warning(f"**{w['product'][:50]}...** (SKU: {w['sku']})  \n"
                                  f"Will stockout on Day {w['stockout_day']} "
                                  f"(Deficit: {w['deficit']:.0f} units)")
        else:
            st.success("No stockouts predicted within the planning period!")
    
    with tab3:
        st.subheader("Location Summary")
        
        # Location metrics
        location_summary = []
        for location in sorted(location_daily_sales.keys()):
            if location.upper() != 'YSXA':
                priority = location_priorities.get(location, 'Pri-4')
                daily_sales = location_daily_sales.get(location, 0)
                
                # Count SKUs needing shipment
                skus_needing = 0
                total_units_needed = 0
                for sku, locs in shipment_requirements.items():
                    if location in locs and locs[location]['required_shipment'] > 0:
                        skus_needing += 1
                        total_units_needed += locs[location]['required_shipment']
                
                location_summary.append({
                    'Location': location,
                    'Priority': priority,
                    'Daily Sales': round(daily_sales, 0),
                    'SKUs Needing Shipment': skus_needing,
                    'Total Units Needed': total_units_needed
                })
        
        df_location_summary = pd.DataFrame(location_summary)
        st.dataframe(df_location_summary, use_container_width=True)
    
    with tab4:
        st.subheader("Download Results")
        
        # Generate Excel file
        excel_output = generate_excel_output(
            shipment_requirements,
            location_priorities,
            inventory_data,
            stockout_warnings,
            target_stock_days,
            inventory_plan_template
        )
        
        # Download button
        st.download_button(
            label="📥 Download Inventory Plan (Excel)",
            data=excel_output,
            file_name=f"Inventory_Plan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        st.info("The Excel file contains the following sheets:\n"
                "- **Shipment_Plan**: Main shipment requirements\n"
                "- **Calculation_Details**: Detailed calculations for each SKU-Location\n"
                "- **Summary**: Overview metrics\n"
                "- **Stockout_Warnings**: Critical alerts (if any)")

else:
    # Instructions
    st.info("👈 Please upload all required files in the sidebar and click 'Process Files' to begin.")
    
    st.markdown("""
    ### Required Files:
    
    1. **Monthly Sales Report** - Should contain:
       - SKU column
       - Product names
       - State-wise sales data
    
    2. **States Mapping File** - Should contain sheets:
       - StatesMapping
       - FC-Location
       - DeliveryTime
    
    3. **Daily Stock Report** - Should contain sheets:
       - Overall stock (with hierarchical structure)
       - In-transit
       - At the FC (optional)
    
    4. **Inventory Plan Template** (Optional) - For filling calculated values
    
    ### How it works:
    
    This tool calculates optimal shipment quantities for each SKU-Location combination based on:
    - Current stock levels (including At FC stock)
    - Daily sales rates
    - Lead times for delivery
    - In-transit inventory
    - Target stock days
    
    The calculation ensures you maintain the specified days of stock at each location while considering lead times and in-transit inventory.
    """)